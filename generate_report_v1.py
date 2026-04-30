"""
Monthly Social Report Generator — v1 (draft, using existing data only)
Generates: March_Social_Report_v1.xlsx
Sheets:  1) Summary (platform columns, March vs Feb where available)
         2) Post Analytics (all March posts, per-platform metrics)
         3) Top Posts (Top 5 per platform by impressions)
"""

import json, os
from datetime import datetime, timezone
from collections import defaultdict

# ── Load data ────────────────────────────────────────────────────────────────

with open("history_march.json") as f:
    history_march = json.load(f).get("history", [])

with open("analytics_results_full.json") as f:
    analytics = json.load(f)   # 29 items

with open("recent_post_analytics.json") as f:
    recent_analytics = json.load(f)   # 10 items (subset of above, same structure)

with open("social_mar.json") as f:
    social_mar = json.load(f)

with open("social_feb.json") as f:
    social_feb = json.load(f)

# ── Build lookup: post_id → analytics data ────────────────────────────────────

# Merge analytics_results_full + recent_post_analytics (dedupe by post_id)
analytics_by_post = {}
for item in analytics + recent_analytics:
    pid = item.get("post_id")
    if pid and pid not in analytics_by_post and item.get("status") == "success":
        analytics_by_post[pid] = item["data"]

print(f"Unique posts with analytics: {len(analytics_by_post)}")

# ── Build lookup: post_id → history record ────────────────────────────────────

history_by_id = {p["id"]: p for p in history_march if p.get("status") == "success"}

# ── Platform list ─────────────────────────────────────────────────────────────

PLATFORMS = ["facebook", "instagram", "linkedin", "youtube", "twitter", "tiktok"]

# ── Section A: Aggregate post counts per platform ────────────────────────────

posts_by_platform = defaultdict(list)
for p in history_march:
    if p.get("status") != "success":
        continue
    for pl in p.get("platforms", []):
        posts_by_platform[pl].append(p)

# ── Section B: Per-post analytics merged with history ────────────────────────

all_rows = []
for platform in PLATFORMS:
    for post in posts_by_platform.get(platform, []):
        pid = post["id"]
        analytics_data = analytics_by_post.get(pid, {}).get(platform, {}).get("analytics", {})
        row = {
            "platform": platform,
            "post_id": pid,
            "post_text": post.get("post", "")[:120],
            "created": post.get("created", ""),
            "impressions": analytics_data.get("impressionsUnique")
                         or analytics_data.get("impressionCount")
                         or analytics_data.get("impressions")
                         or analytics_data.get("pagePostsImpressionsUnique")
                         or analytics_data.get("reachCount")
                         or None,
            "likes": (analytics_data.get("likeCount") or 0)
                   + (analytics_data.get("reactions", {}).get("like", 0))
                   + (analytics_data.get("reactions", {}).get("praise", 0)),
            "comments": analytics_data.get("commentsCount") or analytics_data.get("commentCount") or analytics_data.get("commentCountTotal") or analytics_data.get("comments", 0),
            "shares": analytics_data.get("shareCount") or analytics_data.get("shares", 0),
            "clicks": analytics_data.get("clickCount") or analytics_data.get("clicks") or analytics_data.get("videoViews") or analytics_data.get("pageVideoViews") or None,
            "engagement": analytics_data.get("engagement") or None,
            "video_views": analytics_data.get("videoViews") or analytics_data.get("video_views") or analytics_data.get("viewCountPeriod") or None,
            "has_analytics": bool(analytics_data),
        }
        all_rows.append(row)

print(f"Total rows (post × platform): {len(all_rows)}")

# ── Section C: Social summary (account-level) ────────────────────────────────

def get_social_metric(social_data, platform, key):
    """Extract metric from social endpoint response."""
    info = social_data.get(platform, {})
    a = info.get("analytics", {})
    if platform == "facebook":
        return a.get(key)
    elif platform == "instagram":
        return a.get(key)
    elif platform == "linkedin":
        if key == "followers":
            return a.get("followers", {}).get("totalFollowerCount")
        return a.get(key)
    elif platform == "youtube":
        return a.get(key)
    elif platform == "twitter":
        return a.get(key)
    elif platform == "tiktok":
        return a.get(key)
    return None

# ── Build Top-5 per platform ─────────────────────────────────────────────────

top5_by_platform = defaultdict(list)
for row in all_rows:
    imp = row["impressions"] if row["has_analytics"] else None
    if imp is not None and imp > 0:
        top5_by_platform[row["platform"]].append(row)

for pl in PLATFORMS:
    top5_by_platform[pl] = sorted(
        top5_by_platform[pl],
        key=lambda x: x["impressions"] or 0,
        reverse=True
    )[:5]

# ── Write Excel ──────────────────────────────────────────────────────────────

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("openpyxl not installed; installing...")

if not HAS_OPENPYXL:
    import subprocess
    subprocess.run(["pip", "install", "openpyxl"], check=True)
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ─── Colour palette ───────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1F3864")   # dark navy
SUBHDR_FILL   = PatternFill("solid", fgColor="2E75B6")   # mid blue
ROW_ALT_FILL  = PatternFill("solid", fgColor="DDEEFF")   # light blue tint
GREEN_FILL    = PatternFill("solid", fgColor="E2EFDA")   # light green (MoM +)
RED_FILL      = PatternFill("solid", fgColor="FCE4D6")   # light red (MoM -)
WHITE_FILL    = PatternFill("solid", fgColor="FFFFFF")
GOLD_FILL     = PatternFill("solid", fgColor="FFD700")

HEADER_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TITLE_FONT    = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
BODY_FONT     = Font(name="Calibri", size=10)
BOLD_FONT     = Font(name="Calibri", bold=True, size=10)
CENTER        = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT          = Alignment(horizontal="left",   vertical="center", wrap_text=True)

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def style_header(cell, fill=None):
    cell.font   = HEADER_FONT
    cell.fill   = fill or HEADER_FILL
    cell.alignment = CENTER
    cell.border = thin_border()

def set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — SUMMARY
# ═══════════════════════════════════════════════════════════════════════════════

ws1 = wb.active
ws1.title = "Summary"
ws1.sheet_view.showGridLines = False

# Title row
ws1.merge_cells("A1:M1")
title_cell = ws1["A1"]
title_cell.value = "  Boolell Advisory Mauritius — March 2026 Social Media Report"
title_cell.font  = TITLE_FONT
title_cell.fill  = PatternFill("solid", fgColor="0D1B2A")
title_cell.alignment = LEFT
ws1.row_dimensions[1].height = 36

ws1.merge_cells("A2:M2")
ws1["A2"].value = "  Data coverage: 98 successful posts in March | Analytics available for 29 posts (shown with metrics) | 69 posts marked N/A (analytics pending)"
ws1["A2"].font  = Font(name="Calibri", italic=True, size=9, color="595959")
ws1["A2"].alignment = LEFT
ws1.row_dimensions[2].height = 18

# ── Table 1: Platform Performance ─────────────────────────────────────────────

row = 4
ws1.merge_cells(f"A{row}:M{row}")
hdr = ws1[f"A{row}"]
hdr.value = "  Platform Performance — March 2026 vs February 2026"
hdr.font  = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
hdr.fill  = SUBHDR_FILL
hdr.alignment = LEFT
ws1.row_dimensions[row].height = 24

row += 1
headers1 = [
    "Platform", "Posts\n(Mar)", "Followers\n(Mar)", "Followers\n(Feb)",
    "Follower\nChange", "Impressions\n(Mar)", "Impressions\n(Feb)", "Impression\nChange",
    "Clicks\n(Mar)", "Comments\n(Mar)", "Likes\n(Mar)", "Shares\n(Mar)", "Status"
]
for ci, h in enumerate(headers1, 1):
    c = ws1.cell(row=row, column=ci, value=h)
    style_header(c)
    c.border = thin_border()
ws1.row_dimensions[row].height = 40

# Metric key mapping per platform from social endpoint
SOCIAL_KEYS = {
    "facebook":   {"followers": "followersCount",     "impressions": "pagePostsImpressionsUnique"},
    "instagram":  {"followers": "followersCount",     "impressions": "reachCount"},
    "linkedin":   {"followers": "totalFollowerCount", "impressions": "impressionCount"},
    "youtube":    {"followers": "subscriberCount",    "impressions": "viewCount"},
    "twitter":    {"followers": "followersCount",      "impressions": "tweetCount"},
    "tiktok":     {"followers": "followerCount",      "impressions": "viewCountTotal"},
}

PLATFORM_ICONS = {
    "facebook": "📘", "instagram": "📷", "linkedin": "💼",
    "youtube": "▶️", "twitter": "🐦", "tiktok": "🎵"
}

row += 1
for pi, platform in enumerate(PLATFORMS):
    fill = ROW_ALT_FILL if pi % 2 == 0 else WHITE_FILL

    mar_followers = get_social_metric(social_mar, platform, SOCIAL_KEYS[platform]["followers"]) or 0
    feb_followers = get_social_metric(social_feb, platform, SOCIAL_KEYS[platform]["followers"]) or 0
    follower_delta = mar_followers - feb_followers

    mar_imp = get_social_metric(social_mar, platform, SOCIAL_KEYS[platform]["impressions"]) or 0
    feb_imp = get_social_metric(social_feb, platform, SOCIAL_KEYS[platform]["impressions"]) or 0
    imp_delta = mar_imp - feb_imp

    # Aggregate per-post metrics for this platform
    plat_rows = [r for r in all_rows if r["platform"] == platform]
    post_count = len(plat_rows)
    clicks_sum  = sum(r["clicks"] or 0   for r in plat_rows if r["has_analytics"])
    comments_sum = sum(r["comments"] or 0 for r in plat_rows if r["has_analytics"])
    likes_sum   = sum(r["likes"] or 0     for r in plat_rows if r["has_analytics"])
    shares_sum  = sum(r["shares"] or 0    for r in plat_rows if r["has_analytics"])

    vals = [
        f"{PLATFORM_ICONS[platform]} {platform.capitalize()}",
        post_count,
        mar_followers,
        feb_followers,
        follower_delta,
        mar_imp,
        feb_imp,
        imp_delta,
        clicks_sum,
        comments_sum,
        likes_sum,
        shares_sum,
        "✓ Draft"
    ]
    for ci, v in enumerate(vals, 1):
        c = ws1.cell(row=row, column=ci, value=v)
        c.font   = BOLD_FONT if ci == 1 else BODY_FONT
        c.fill   = fill
        c.alignment = CENTER if ci > 1 else LEFT
        c.border = thin_border()
        # Colour follower/impression delta
        if ci == 5 and isinstance(v, (int, float)) and v > 0:
            c.fill = GREEN_FILL
        elif ci == 5 and isinstance(v, (int, float)) and v < 0:
            c.fill = RED_FILL
        if ci == 8 and isinstance(v, (int, float)) and v > 0:
            c.fill = GREEN_FILL
        elif ci == 8 and isinstance(v, (int, float)) and v < 0:
            c.fill = RED_FILL
    ws1.row_dimensions[row].height = 22
    row += 1

# ── Column widths for Summary ────────────────────────────────────────────────
col_widths = [22, 10, 14, 14, 12, 14, 14, 12, 10, 10, 10, 10, 12]
for ci, w in enumerate(col_widths, 1):
    set_col_width(ws1, get_column_letter(ci), w)

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — POST ANALYTICS
# ═══════════════════════════════════════════════════════════════════════════════

ws2 = wb.create_sheet("Post Analytics")
ws2.sheet_view.showGridLines = False

ws2.merge_cells("A1:K1")
t = ws2["A1"]
t.value = "  March 2026 — All Posts with Analytics"
t.font  = TITLE_FONT
t.fill  = PatternFill("solid", fgColor="0D1B2A")
t.alignment = LEFT
ws2.row_dimensions[1].height = 36

headers2 = [
    "Platform", "Post ID", "Post Text (truncated)", "Created (UTC)",
    "Impressions", "Likes", "Comments", "Shares", "Clicks", "Video Views", "Has Analytics"
]
for ci, h in enumerate(headers2, 1):
    c = ws2.cell(row=2, column=ci, value=h)
    style_header(c)
ws2.row_dimensions[2].height = 36

for ri, row_data in enumerate(all_rows, 3):
    fill = ROW_ALT_FILL if ri % 2 == 0 else WHITE_FILL
    vals2 = [
        row_data["platform"].capitalize(),
        row_data["post_id"],
        row_data["post_text"],
        row_data["created"][:10],
        row_data["impressions"] if row_data["has_analytics"] else "N/A",
        row_data["likes"]       if row_data["has_analytics"] else "N/A",
        row_data["comments"]     if row_data["has_analytics"] else "N/A",
        row_data["shares"]      if row_data["has_analytics"] else "N/A",
        row_data["clicks"]      if row_data["has_analytics"] else "N/A",
        row_data["video_views"] if row_data["has_analytics"] else "N/A",
        "✓" if row_data["has_analytics"] else "⚠ N/A",
    ]
    for ci, v in enumerate(vals2, 1):
        c = ws2.cell(row=ri, column=ci, value=v)
        c.font = BODY_FONT
        c.fill = fill
        c.alignment = LEFT if ci == 3 else CENTER
        c.border = thin_border()
        if ci == 11 and v == "⚠ N/A":
            c.fill = PatternFill("solid", fgColor="FFF2CC")
    ws2.row_dimensions[ri].height = 18

col_widths2 = [14, 30, 55, 16, 14, 10, 10, 10, 10, 12, 14]
for ci, w in enumerate(col_widths2, 1):
    set_col_width(ws2, get_column_letter(ci), w)

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — TOP POSTS
# ═══════════════════════════════════════════════════════════════════════════════

ws3 = wb.create_sheet("Top Posts")
ws3.sheet_view.showGridLines = False

ws3.merge_cells("A1:J1")
t3 = ws3["A1"]
t3.value = "  Top 5 Posts per Platform — March 2026 (ranked by Impressions)"
t3.font  = TITLE_FONT
t3.fill  = PatternFill("solid", fgColor="0D1B2A")
t3.alignment = LEFT
ws3.row_dimensions[1].height = 36

headers3 = [
    "Rank", "Platform", "Post Title / Opening Text",
    "Impressions", "Likes", "Comments", "Shares", "Clicks", "Video Views", "Data Status"
]
for ci, h in enumerate(headers3, 1):
    c = ws3.cell(row=2, column=ci, value=h)
    style_header(c)
ws3.row_dimensions[2].height = 36

rank_colours = {
    1: PatternFill("solid", fgColor="FFD700"),   # Gold
    2: PatternFill("solid", fgColor="C0C0C0"),   # Silver
    3: PatternFill("solid", fgColor="CD7F32"),   # Bronze
    4: PatternFill("solid", fgColor="DDEEFF"),
    5: PatternFill("solid", fgColor="EAF2FF"),
}

row = 3
for platform in PLATFORMS:
    # Section header per platform
    ws3.merge_cells(f"A{row}:J{row}")
    ph = ws3[f"A{row}"]
    ph.value = f"  {PLATFORM_ICONS.get(platform, '•')}  {platform.upper()}"
    ph.font  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    ph.fill  = SUBHDR_FILL
    ph.alignment = LEFT
    ws3.row_dimensions[row].height = 22
    row += 1

    top5 = top5_by_platform.get(platform, [])
    if not top5:
        ws3.merge_cells(f"A{row}:J{row}")
        nc = ws3[f"A{row}"]
        nc.value = "  No posts with analytics available for this platform."
        nc.font  = Font(name="Calibri", italic=True, color="595959")
        nc.fill  = WHITE_FILL
        nc.alignment = LEFT
        ws3.row_dimensions[row].height = 18
        row += 1
        continue

    for rank, trow in enumerate(top5, 1):
        rfill = rank_colours.get(rank, WHITE_FILL)
        # Truncate post text to first 100 chars for title
        post_title = (trow["post_text"] or "No text")[:100]
        vals3 = [
            f"#{rank}",
            trow["platform"].capitalize(),
            post_title,
            trow["impressions"],
            trow["likes"],
            trow["comments"],
            trow["shares"],
            trow["clicks"],
            trow["video_views"],
            "✓ Data",
        ]
        for ci, v in enumerate(vals3, 1):
            c = ws3.cell(row=row, column=ci, value=v)
            c.font = BOLD_FONT if ci == 1 else BODY_FONT
            c.fill = rfill
            c.alignment = CENTER if ci in (1, 2, 10) else LEFT
            c.border = thin_border()
        ws3.row_dimensions[row].height = 36
        row += 1

col_widths3 = [8, 14, 50, 14, 10, 10, 10, 10, 12, 12]
for ci, w in enumerate(col_widths3, 1):
    set_col_width(ws3, get_column_letter(ci), w)

# ── Save ──────────────────────────────────────────────────────────────────────

out_path = "March_Social_Report_v1.xlsx"
wb.save(out_path)
print(f"\n✅  Saved: {out_path}")
print(f"    Sheet 1 'Summary'        — platform performance table")
print(f"    Sheet 2 'Post Analytics' — {len(all_rows)} post rows")
total_top = sum(len(v) for v in top5_by_platform.values())
print(f"    Sheet 3 'Top Posts'      — {total_top} top posts across {len([v for v in top5_by_platform.values() if v])} platforms")
