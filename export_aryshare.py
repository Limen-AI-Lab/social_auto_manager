#!/usr/bin/env python3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "AryShare Posts Mar 26-27"

# Styles
header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
header_fill = PatternFill(start_color='2E4057', end_color='2E4057', fill_type='solid')
plat_fills = {
    'Facebook':  PatternFill(start_color='D6EAF8', end_color='D6EAF8', fill_type='solid'),
    'TikTok':    PatternFill(start_color='FADBD8', end_color='FADBD8', fill_type='solid'),
    'Twitter/X': PatternFill(start_color='D5F5E3', end_color='D5F5E3', fill_type='solid'),
    'YouTube':   PatternFill(start_color='FCF3CF', end_color='FCF3CF', fill_type='solid'),
    'Instagram': PatternFill(start_color='F5EEF8', end_color='F5EEF8', fill_type='solid'),
    'LinkedIn':  PatternFill(start_color='E8DAEF', end_color='E8DAEF', fill_type='solid'),
}
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# Headers (row 1)
headers = ['Date', 'Topic', 'Platform', 'Post URL', 'Views', 'Likes', 'Comments', 'Shares', 'Reach', 'Impressions', 'Engagement']
ws.append(headers)
for col_idx, cell in enumerate(ws[1], 1):
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = thin_border

# Column widths
col_widths = [12, 40, 12, 50, 12, 10, 12, 10, 12, 14, 14]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Data rows - 2 topics, 2 days, 6 platforms each = 12 posts
posts = [
    # ===================== TOPIC 1: Hidden Fees & Poor Communication (March 27) =====================
    {
        'date': '2026-03-27', 'topic': 'Hidden Fees & Poor Communication',
        'platform': 'Facebook',
        'url': 'https://www.facebook.com/watch/?v=952665770791879',
        'views': 116, 'likes': 0, 'comments': 0, 'shares': 0,
        'reach': 107, 'impressions': 107, 'engagement': 0
    },
    {
        'date': '2026-03-27', 'topic': 'Hidden Fees & Poor Communication',
        'platform': 'TikTok',
        'url': 'https://www.tiktok.com/@boolelladvisorymauritius/video/7621989841894493458',
        'views': 299, 'likes': 2, 'comments': 0, 'shares': 0,
        'reach': 29, 'impressions': 299, 'engagement': 4
    },
    {
        'date': '2026-03-27', 'topic': 'Hidden Fees & Poor Communication',
        'platform': 'Twitter/X',
        'url': 'https://twitter.com/boolelladvisory/status/2037583723575521424',
        'views': 1, 'likes': 0, 'comments': 0, 'shares': 0,
        'reach': 12, 'impressions': 12, 'engagement': 0
    },
    {
        'date': '2026-03-27', 'topic': 'Hidden Fees & Poor Communication',
        'platform': 'YouTube',
        'url': 'https://youtu.be/9sjEAbAnaHA',
        'views': 1, 'likes': 0, 'comments': 0, 'shares': 0,
        'reach': 1, 'impressions': 1, 'engagement': 0
    },
    {
        'date': '2026-03-27', 'topic': 'Hidden Fees & Poor Communication',
        'platform': 'Instagram',
        'url': 'https://www.instagram.com/reel/DWZXedpjcpY/',
        'views': 44, 'likes': 2, 'comments': 0, 'shares': 1,
        'reach': 29, 'impressions': 44, 'engagement': 3
    },
    {
        'date': '2026-03-27', 'topic': 'Hidden Fees & Poor Communication',
        'platform': 'LinkedIn',
        'url': 'https://www.linkedin.com/feed/update/urn:li:ugcPost:7443348955041538048',
        'views': 13, 'likes': 0, 'comments': 0, 'shares': 0,
        'reach': 25, 'impressions': 41, 'engagement': 1
    },
    # ===================== TOPIC 2: CSP Transition (March 26) =====================
    {
        'date': '2026-03-26', 'topic': 'CSP Transition - Is It Time to Change?',
        'platform': 'Facebook',
        'url': 'https://www.facebook.com/watch/?v=4229791830609807',
        'views': 6, 'likes': 0, 'comments': 0, 'shares': 0,
        'reach': 6, 'impressions': 6, 'engagement': 0
    },
    {
        'date': '2026-03-26', 'topic': 'CSP Transition - Is It Time to Change?',
        'platform': 'TikTok',
        'url': 'https://www.tiktok.com/@boolelladvisorymauritius/video/7621483050718842132',
        'views': 1095, 'likes': 15, 'comments': 0, 'shares': 2,
        'reach': 911, 'impressions': 1095, 'engagement': 21
    },
    {
        'date': '2026-03-26', 'topic': 'CSP Transition - Is It Time to Change?',
        'platform': 'Twitter/X',
        'url': 'https://twitter.com/boolelladvisory/status/2037088841488531860',
        'views': 4, 'likes': 0, 'comments': 0, 'shares': 0,
        'reach': 5, 'impressions': 5, 'engagement': 0
    },
    {
        'date': '2026-03-26', 'topic': 'CSP Transition - Is It Time to Change?',
        'platform': 'YouTube',
        'url': 'https://youtu.be/4bqS6YBeQNI',
        'views': 85, 'likes': 1, 'comments': 0, 'shares': 2,
        'reach': 85, 'impressions': 85, 'engagement': 3
    },
    {
        'date': '2026-03-26', 'topic': 'CSP Transition - Is It Time to Change?',
        'platform': 'Instagram',
        'url': 'https://www.instagram.com/reel/DWV2hMJCC_n/',
        'views': 33, 'likes': 1, 'comments': 0, 'shares': 0,
        'reach': 24, 'impressions': 33, 'engagement': 1
    },
    {
        'date': '2026-03-26', 'topic': 'CSP Transition - Is It Time to Change?',
        'platform': 'LinkedIn',
        'url': 'https://www.linkedin.com/feed/update/urn:li:ugcPost:7442854254932910080',
        'views': 30, 'likes': 2, 'comments': 2, 'shares': 1,
        'reach': 62, 'impressions': 109, 'engagement': 11
    },
]

for row_idx, p in enumerate(posts, 2):
    row = [p['date'], p['topic'], p['platform'], p['url'],
           p['views'], p['likes'], p['comments'], p['shares'],
           p['reach'], p['impressions'], p['engagement']]
    ws.append(row)
    fill = plat_fills[p['platform']]
    for cell in ws[row_idx]:
        cell.fill = fill
        cell.border = thin_border
        cell.alignment = center if cell.column != 4 else left

# Freeze top row
ws.freeze_panes = 'A2'

# Save
out_path = '/Users/huaweiwei/Desktop/截图/SAMA---Social-Auto-Manager-main/AryShare_Posts_Mar26-27_Data.xlsx'
wb.save(out_path)
print(f'Saved to {out_path}')
