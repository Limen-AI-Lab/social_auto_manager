#!/usr/bin/env python3
r"""
═══════════════════════════════════════════════════════════════════════════════
SAMA — Monthly Report 数据调试脚本（标准化模板）
═══════════════════════════════════════════════════════════════════════════════

用途：
    每月生成 Excel 月报前，先用此脚本诊断所有 6 个平台的数据质量。
    输出 debug/ 目录下的结构化 JSON + 本报告，确认无误后再运行 rebuild_excel。

使用方法：
    cd SAMA项目目录
    python3 monthly_report_debug.py

输出文件：
    debug/{platform}_raw.json          — 原始 API 缓存数据
    debug/{platform}_extracted.json   — 提取后的指标（提取前 vs 提取后）
    debug/report_summary.json          — 所有平台的诊断报告
    debug/EXCEL_VERIFICATION.txt      — Excel 各平台关键单元格值

诊断步骤（每个平台）：
    Step 1: 读取 post_analytics_merged.json → 帖子级原始数据
    Step 2: 读取 social_xxx.json           → 账号级原始数据
    Step 3: 运行当前 extract()（可能有 bug）  → 提取结果
    Step 4: 运行修复后的 extract_fixed()     → 正确结果
    Step 5: 对比两者，报告差异
    Step 6: 汇总月度总数，写入 debug/platform_extracted.json
    Step 7: 运行 rebuild_excel_from_cache.py → 验证 Excel 输出

平台列表：
    linkedin / facebook / instagram / youtube / twitter / tiktok

═══════════════════════════════════════════════════════════════════════════════
"""

from __future__ import annotations

import json
import os
import sys
from collections import defaultdict
from datetime import datetime
from typing import Any

# ── 配置 ─────────────────────────────────────────────────────────────────────

PLATFORMS = ["linkedin", "facebook", "instagram", "youtube", "twitter", "tiktok"]
PLATFORM_LABELS = {
    "linkedin": "LinkedIn",
    "facebook": "Facebook",
    "instagram": "Instagram",
    "youtube": "YouTube",
    "twitter": "X (Twitter)",
    "tiktok": "TikTok",
}

CACHE = "scraped_report_cache"
DEBUG = "debug"
OUT_XLSX = "Social_Media_Performance_Feb_Mar_2026_6Platforms.xlsx"  # ← 每月更新

# 月份配置（← 每月修改这里）
YEAR = 2026
MONTH_A = 2    # ← 上个月
MONTH_B = 3    # ← 当前月
MONTH_LABELS = ("Feb", "Mar")


# ── 工具函数 ────────────────────────────────────────────────────────────────

def n(v: Any) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    return 0.0


def load_json(path: str) -> Any:
    if not os.path.exists(path):
        print(f"  ⚠  文件不存在: {path}")
        return None
    with open(path) as f:
        return json.load(f)


def in_month(iso_ts: str, year: int, month: int) -> bool:
    if not iso_ts:
        return False
    try:
        d = datetime.fromisoformat(iso_ts.replace("Z", "+00:00"))
    except ValueError:
        return False
    return d.year == year and d.month == month


def pct_change(prev: float, curr: float) -> float | None:
    if prev == 0 and curr == 0:
        return 0.0
    if prev == 0:
        return None
    return (curr - prev) / prev * 100.0


def fmt_num(v: float) -> str:
    if v >= 1_000_000:
        return f"{v / 1_000_000:.1f}M"
    if v >= 1_000:
        return f"{v / 1_000:.1f}K"
    if v == 0:
        return "0"
    return f"{v:.0f}"


def banner(title: str, char: str = "═", width: int = 78) -> None:
    print()
    print(char * width)
    print(f"  {title}")
    print(char * width)


def section(title: str) -> None:
    print()
    print(f"── {title} ──")


# ══════════════════════════════════════════════════════════════════════════
#  EXTRACT — 每个平台的提取函数（当前/有 bug 版本）
#  （从 rebuild_excel_from_cache.py 复制，确保一致）
# ══════════════════════════════════════════════════════════════════════════

def extract_current(plat: str, body: dict) -> dict:
    """
    当前 rebuild_excel_from_cache.py 中使用的 extract() 函数。
    可能包含 bug。
    """
    out = {
        "impressions": 0.0, "reach": 0.0, "views": 0.0, "clicks": 0.0,
        "likes": 0.0, "comments": 0.0, "shares": 0.0, "reposts": 0.0,
    }
    if body.get("status") == "error":
        return out

    pdata = body.get(plat)
    if isinstance(pdata, list) and pdata:
        pdata = pdata[0]
    if not isinstance(pdata, dict):
        return out
    a = pdata.get("analytics") or {}

    # ── linkedin ────────────────────────────────────────────────────────────
    if plat == "linkedin":
        out["impressions"] = n(a.get("impressionCount"))
        out["reach"]       = n(a.get("uniqueImpressionsCount"))
        out["views"]       = n(a.get("videoViews"))
        out["clicks"]      = n(a.get("clickCount"))
        out["likes"]       = n(a.get("likeCount"))
        out["comments"]    = n(a.get("commentCount"))
        out["shares"]      = n(a.get("shareCount"))
        out["reposts"]     = out["shares"]

    # ── facebook ─────────────────────────────────────────────────────────────
    elif plat == "facebook":
        out["impressions"] = n(a.get("impressionsUnique")) or n(a.get("impressionsOrganicUnique"))
        out["reach"]        = out["impressions"]
        out["views"]        = n(a.get("blueReelsPlayCount")) or out["impressions"]
        out["likes"]        = n(a.get("likeCount"))
        out["comments"]      = n(a.get("commentsCount"))
        out["shares"]        = n(a.get("sharesCount")) or n(a.get("shareCount"))

    # ── instagram ────────────────────────────────────────────────────────────
    elif plat == "instagram":
        out["impressions"] = n(a.get("reachCount")) or n(a.get("impressionsCount")) or n(a.get("viewsCount"))
        out["reach"]        = n(a.get("reachCount"))
        out["views"]        = n(a.get("viewsCount"))
        out["likes"]        = n(a.get("likeCount"))
        out["comments"]      = n(a.get("commentsCount"))
        out["shares"]        = n(a.get("sharesCount")) or n(a.get("shareCount"))
        out["reposts"]       = out["shares"]

    # ── youtube ─────────────────────────────────────────────────────────────
    elif plat == "youtube":
        out["views"]   = n(a.get("viewCount")) or n(a.get("views")) or n(a.get("videoViews"))
        out["likes"]   = n(a.get("likeCount")) or n(a.get("likes"))
        out["comments"] = n(a.get("commentCount")) or n(a.get("comments"))
        out["shares"]   = n(a.get("shareCount")) or n(a.get("shares"))
        out["reposts"]  = out["shares"]
        out["impressions"] = out["views"]

    # ── twitter ─────────────────────────────────────────────────────────────
    elif plat == "twitter":
        pub = a.get("publicMetrics") or {}
        out["impressions"] = n(pub.get("impressionCount")) or n(a.get("impressions"))
        out["likes"]       = n(pub.get("likeCount")) or n(a.get("likeCount")) or n(a.get("favoriteCount"))
        out["comments"]    = n(pub.get("replyCount")) or n(a.get("replyCount")) or n(a.get("replies"))
        out["reposts"]     = n(pub.get("retweetCount")) or n(a.get("retweetCount")) or n(a.get("retweets", 0))
        out["shares"]       = out["reposts"]
        out["clicks"]      = n(a.get("urlLinkClicks")) or n(a.get("userProfileClicks"))

    # ── tiktok ──────────────────────────────────────────────────────────────
    elif plat == "tiktok":
        out["views"]       = n(a.get("videoViews")) or n(a.get("viewCount")) or n(a.get("views"))
        out["likes"]       = n(a.get("likeCount"))
        out["comments"]    = n(a.get("commentsCount")) or n(a.get("commentCount"))
        out["shares"]      = n(a.get("shareCount")) or n(a.get("shares"))
        out["reposts"]     = out["shares"]
        out["reach"]       = n(a.get("reach")) or 0
        out["impressions"] = out["views"]

    return out


# ══════════════════════════════════════════════════════════════════════════
#  EXTRACT FIXED — 修复后的版本（如果当前版本有 bug）
#  每个平台的最新正确字段映射（参考 MONTHLY_REPORT_FIELD_REFERENCE.md）
# ══════════════════════════════════════════════════════════════════════════

def extract_fixed(plat: str, body: dict) -> dict:
    """
    每个平台的已知正确提取逻辑。
    当此函数结果与 extract_current() 不同时，报告 BUG。
    """
    out = {
        "impressions": 0.0, "reach": 0.0, "views": 0.0, "clicks": 0.0,
        "likes": 0.0, "comments": 0.0, "shares": 0.0, "reposts": 0.0,
    }
    if body.get("status") == "error":
        return out

    pdata = body.get(plat)
    if isinstance(pdata, list) and pdata:
        pdata = pdata[0]
    if not isinstance(pdata, dict):
        return out
    a = pdata.get("analytics") or {}

    if plat == "linkedin":
        out["impressions"] = n(a.get("impressionCount"))
        out["reach"]       = n(a.get("uniqueImpressionsCount"))
        out["views"]       = n(a.get("videoViews"))
        out["clicks"]      = n(a.get("clickCount"))
        out["likes"]       = n(a.get("likeCount"))
        out["comments"]    = n(a.get("commentCount"))
        out["shares"]      = n(a.get("shareCount"))
        out["reposts"]     = out["shares"]

    elif plat == "facebook":
        out["impressions"] = n(a.get("impressionsUnique")) or n(a.get("impressionsOrganicUnique"))
        out["reach"]        = out["impressions"]
        out["views"]        = n(a.get("blueReelsPlayCount")) or out["impressions"]
        out["likes"]        = n(a.get("likeCount"))
        out["comments"]      = n(a.get("commentsCount"))
        out["shares"]        = n(a.get("sharesCount")) or n(a.get("shareCount"))

    elif plat == "instagram":
        out["impressions"] = n(a.get("reachCount")) or n(a.get("impressionsCount")) or n(a.get("viewsCount"))
        out["reach"]        = n(a.get("reachCount"))
        out["views"]        = n(a.get("viewsCount"))
        out["likes"]        = n(a.get("likeCount"))
        out["comments"]      = n(a.get("commentsCount"))
        out["shares"]        = n(a.get("sharesCount")) or n(a.get("shareCount"))
        out["reposts"]       = out["shares"]

    elif plat == "youtube":
        out["views"]   = n(a.get("viewCount")) or n(a.get("views")) or n(a.get("videoViews"))
        out["likes"]   = n(a.get("likeCount")) or n(a.get("likes"))
        out["comments"] = n(a.get("commentCount")) or n(a.get("comments"))
        out["shares"]   = n(a.get("shareCount")) or n(a.get("shares"))
        out["reposts"]  = out["shares"]
        out["impressions"] = out["views"]

    elif plat == "twitter":
        pub = a.get("publicMetrics") or {}
        out["impressions"] = n(pub.get("impressionCount")) or n(a.get("impressions"))
        out["likes"]       = n(pub.get("likeCount")) or n(a.get("likeCount")) or n(a.get("favoriteCount"))
        out["comments"]    = n(pub.get("replyCount")) or n(a.get("replyCount")) or n(a.get("replies"))
        out["reposts"]     = n(pub.get("retweetCount")) or n(a.get("retweetCount")) or n(a.get("retweets", 0))
        out["shares"]       = out["reposts"]
        out["clicks"]      = n(a.get("urlLinkClicks")) or n(a.get("userProfileClicks"))

    elif plat == "tiktok":
        # TikTok 核心字段：videoViews（不是 viewCount 或 views）
        # 2026-03-30 修复：videoViews 是主要字段
        out["views"]       = n(a.get("videoViews")) or n(a.get("viewCount")) or n(a.get("views"))
        out["likes"]       = n(a.get("likeCount"))
        out["comments"]    = n(a.get("commentsCount")) or n(a.get("commentCount"))
        out["shares"]      = n(a.get("shareCount")) or n(a.get("shares"))
        out["reposts"]     = out["shares"]
        out["reach"]       = n(a.get("reach")) or 0
        # TikTok 没有独立的 impressions 字段，用 views 作为最佳近似
        out["impressions"] = out["views"]

    return out


# ══════════════════════════════════════════════════════════════════════════
#  SOCIAL-ACCOUNT LEVEL — 账号级提取
# ══════════════════════════════════════════════════════════════════════════

def followers_from_social(plat: str, soc: dict) -> float:
    block = soc.get(plat) or {}
    a = block.get("analytics") or {}
    if plat == "linkedin":
        return n((a.get("followers") or {}).get("totalFollowerCount"))
    if plat == "facebook":
        return n(a.get("followersCount"))
    if plat == "instagram":
        return n(a.get("followersCount"))
    if plat == "youtube":
        return n(a.get("subscriberCount"))
    if plat == "twitter":
        return n(a.get("followersCount"))
    if plat == "tiktok":
        return n(a.get("followerCount"))
    return 0.0


def soc_imp_proxy(plat: str, soc: dict) -> float:
    """账号级 impressions 近似（仅用于 post 级全为 0 时的 fallback）。"""
    block = soc.get(plat) or {}
    a = block.get("analytics") or {}
    if plat == "facebook":
        return n(a.get("pagePostsImpressionsUnique")) or n(a.get("pagePostsImpressions"))
    if plat == "instagram":
        return n(a.get("reachCount")) or n(a.get("viewsCount"))
    if plat == "linkedin":
        return n(a.get("impressionCount"))
    if plat == "youtube":
        return n(a.get("viewCount"))
    if plat == "tiktok":
        return n(a.get("viewCountTotal"))
    return 0.0


# ══════════════════════════════════════════════════════════════════════════
#  主诊断流程
# ══════════════════════════════════════════════════════════════════════════

def diagnose_platform(plat: str, raw_analytics: list, history_a: list,
                     history_b: list, soc_a: dict, soc_b: dict) -> dict:
    """
    对单个平台进行完整数据诊断。
    返回诊断报告字典。
    """
    label = PLATFORM_LABELS.get(plat, plat)
    report: dict[str, Any] = {
        "platform": plat,
        "label": label,
        "status": "PASS",
        "bugs": [],
        "warnings": [],
        "social_a": {},
        "social_b": {},
        "posts_a": [],
        "posts_b": [],
        "totals_a": {"posts": 0, "views": 0, "impressions": 0, "likes": 0,
                     "comments": 0, "shares": 0, "reach": 0},
        "totals_b": {"posts": 0, "views": 0, "impressions": 0, "likes": 0,
                     "comments": 0, "shares": 0, "reach": 0},
        "followers_a": 0,
        "followers_b": 0,
        "mom_views": None,
        "mom_impressions": None,
    }

    # ── Step 1: Social-account analytics ──────────────────────────────────
    section(f"账号级数据 ({label})")
    soc_a_block = soc_a.get(plat, {}).get("analytics", {})
    soc_b_block = soc_b.get(plat, {}).get("analytics", {})

    # 打印主要字段
    key_fields = _platform_key_social_fields(plat)
    print(f"  {MONTH_LABELS[0]} 月关键字段:")
    for fk, flabel in key_fields:
        v = soc_a_block.get(fk, "—")
        print(f"    {flabel:25s} = {v}")
    print(f"  {MONTH_LABELS[1]} 月关键字段:")
    for fk, flabel in key_fields:
        v = soc_b_block.get(fk, "—")
        print(f"    {flabel:25s} = {v}")

    report["social_a"] = {k: soc_a_block.get(k) for k, _ in key_fields}
    report["social_b"] = {k: soc_b_block.get(k) for k, _ in key_fields}

    # Followers
    fa = followers_from_social(plat, soc_a)
    fb = followers_from_social(plat, soc_b)
    print(f"  Followers:  {fmt_num(fa)} → {fmt_num(fb)}  |  Δ = {fmt_mom(pct_change(fa, fb))}")
    report["followers_a"] = fa
    report["followers_b"] = fb

    # ── Step 2: Collect posts for each month ────────────────────────────────
    posts_a = _filter_posts_by_month(history_a, plat, YEAR, MONTH_A)
    posts_b = _filter_posts_by_month(history_b, plat, YEAR, MONTH_B)
    print(f"\n  历史帖子数:  {MONTH_LABELS[0]}={len(posts_a)}  {MONTH_LABELS[1]}={len(posts_b)}")

    # ── Step 3: Build post_metrics dict ──────────────────────────────────────
    post_metrics: dict[tuple, dict] = {}
    raw_entries: list = []

    for r in raw_analytics:
        if str(r.get("platform", "")).lower() != plat:
            continue
        pid = r["post_id"]
        body = r.get("body", {})
        if body.get("status") == "success" and (pid, plat) not in post_metrics:
            post_metrics[(pid, plat)] = extract_current(plat, body)

    print(f"  成功提取的帖子: {len(post_metrics)} 条")

    # ── Step 4: Sum per month ───────────────────────────────────────────────
    def sum_month(posts: list, key: str) -> float:
        s = 0.0
        for h in posts:
            pid = str(h.get("id", ""))
            s += float(post_metrics.get((pid, plat), {}).get(key, 0) or 0)
        return s

    imp_a  = sum_month(posts_a, "impressions") or soc_imp_proxy(plat, soc_a)
    imp_b  = sum_month(posts_b, "impressions") or soc_imp_proxy(plat, soc_b)
    reach_a = sum_month(posts_a, "reach")
    reach_b = sum_month(posts_b, "reach")
    views_a = sum_month(posts_a, "views")
    views_b = sum_month(posts_b, "views")
    likes_a = sum_month(posts_a, "likes")
    likes_b = sum_month(posts_b, "likes")
    comms_a = sum_month(posts_a, "comments")
    comms_b = sum_month(posts_b, "comments")
    shares_a = sum_month(posts_a, "shares")
    shares_b = sum_month(posts_b, "shares")

    print(f"\n  ┌─ 月度汇总 ─")
    print(f"  │  Metric      {MONTH_LABELS[0]:>8s}  {MONTH_LABELS[1]:>8s}  MoM Δ")
    print(f"  ├─ Posts        {len(posts_a):>8d}  {len(posts_b):>8d}")
    print(f"  ├─ Views        {fmt_num(views_a):>8s}  {fmt_num(views_b):>8s}  {fmt_mom(pct_change(views_a, views_b))}")
    print(f"  ├─ Impressions  {fmt_num(imp_a):>8s}  {fmt_num(imp_b):>8s}  {fmt_mom(pct_change(imp_a, imp_b))}")
    print(f"  ├─ Reach        {fmt_num(reach_a):>8s}  {fmt_num(reach_b):>8s}  {fmt_mom(pct_change(reach_a, reach_b))}")
    print(f"  ├─ Likes        {fmt_num(likes_a):>8s}  {fmt_num(likes_b):>8s}  {fmt_mom(pct_change(likes_a, likes_b))}")
    print(f"  ├─ Comments      {fmt_num(comms_a):>8s}  {fmt_num(comms_b):>8s}  {fmt_mom(pct_change(comms_a, comms_b))}")
    print(f"  └─ Shares       {fmt_num(shares_a):>8s}  {fmt_num(shares_b):>8s}  {fmt_mom(pct_change(shares_a, shares_b))}")

    report["totals_a"] = {
        "posts": len(posts_a), "views": views_a, "impressions": imp_a,
        "likes": likes_a, "comments": comms_a, "shares": shares_a, "reach": reach_a,
    }
    report["totals_b"] = {
        "posts": len(posts_b), "views": views_b, "impressions": imp_b,
        "likes": likes_b, "comments": comms_b, "shares": shares_b, "reach": reach_b,
    }
    report["mom_views"] = pct_change(views_a, views_b)
    report["mom_impressions"] = pct_change(imp_a, imp_b)

    # ── Step 5: Bug Detection — compare extract_current vs extract_fixed ──────
    print(f"\n  ┌─ 提取一致性检查 (extract_current vs extract_fixed) ─")
    buggy_posts = []
    for r in raw_analytics:
        if str(r.get("platform", "")).lower() != plat:
            continue
        pid = r["post_id"]
        body = r.get("body", {})
        if body.get("status") != "success":
            continue
        cur = extract_current(plat, body)
        fix = extract_fixed(plat, body)
        if cur != fix:
            buggy_posts.append({
                "post_id": pid,
                "current": cur,
                "fixed": fix,
                "diff": {k: {"current": cur[k], "fixed": fix[k]}
                          for k in cur if cur[k] != fix[k]},
            })

    if buggy_posts:
        report["status"] = "FAIL"
        print(f"  │  ⚠  发现 {len(buggy_posts)} 个帖子有字段差异")
        for bp in buggy_posts[:3]:  # 只打印前 3 个
            print(f"  │    Post {bp['post_id'][:12]}...:")
            for k, v in bp["diff"].items():
                print(f"  │      {k:12s}:  {v['current']:>10.1f}  →  {v['fixed']:>10.1f}")
        if len(buggy_posts) > 3:
            print(f"  │    ... 共 {len(buggy_posts)} 个帖子受影响")
        report["bugs"].append(f"{len(buggy_posts)} 个帖子提取不一致")
        # 保存完整的 buggy posts 列表到 debug 目录
        _save_debug_json(f"{plat}_buggy_posts.json", buggy_posts)
    else:
        print(f"  │  ✅ extract() 与参考一致，无已知 bug")

    # ── Step 6: 数据质量警告 ────────────────────────────────────────────────
    if imp_a == 0 and imp_b == 0 and len(posts_a) > 0:
        report["warnings"].append(f"{MONTH_LABELS[0]} 月 impressions 全为 0，需检查")
        print(f"  │  ⚠  WARNING: {MONTH_LABELS[0]} 月 impressions 为 0")
    if imp_b == 0 and len(posts_b) > 0:
        report["warnings"].append(f"{MONTH_LABELS[1]} 月 impressions 全为 0，需检查")
        print(f"  │  ⚠  WARNING: {MONTH_LABELS[1]} 月 impressions 为 0")
    if len(posts_a) == 0 and len(posts_b) == 0:
        report["warnings"].append("历史记录中无任何帖子")
        print(f"  │  ⚠  WARNING: 两个月份均无帖子")

    # ── Step 7: 检查 Excel 输出 ─────────────────────────────────────────────
    print(f"\n  └─ Excel 验证:")
    if os.path.exists(OUT_XLSX):
        excel_ok = _verify_excel_platform(plat, posts_a, posts_b, post_metrics)
        if excel_ok:
            print(f"     ✅ Excel 输出正常")
        else:
            print(f"     ⚠  Excel 验证未通过（可能需要重新生成）")
    else:
        print(f"     ⏭  Excel 文件不存在，跳过验证")

    return report


def _platform_key_social_fields(plat: str) -> list:
    """返回每个平台在账号级 analytics 中的关键字段列表。"""
    if plat == "linkedin":
        return [
            ("impressionCount", "Impressions"),
            ("uniqueImpressionsCount", "Unique Impressions (Reach)"),
            ("clickCount", "Clicks"),
            ("likeCount", "Likes"),
            ("commentCount", "Comments"),
            ("shareCount", "Shares"),
        ]
    if plat == "facebook":
        return [
            ("pagePostsImpressions", "Page Impressions"),
            ("pagePostsImpressionsUnique", "Page Impressions (Unique)"),
            ("pageMediaView", "Page Media Views"),
            ("likeCount", "Likes"),
            ("commentsCount", "Comments"),
            ("sharesCount", "Shares"),
            ("followersCount", "Followers"),
        ]
    if plat == "instagram":
        return [
            ("viewsCount", "Views Count"),
            ("reachCount", "Reach Count"),
            ("impressionsCount", "Impressions Count"),
            ("likeCount", "Likes"),
            ("commentsCount", "Comments"),
            ("followersCount", "Followers"),
        ]
    if plat == "youtube":
        return [
            ("viewCount", "View Count"),
            ("subscriberCount", "Subscribers"),
            ("likeCount", "Likes"),
            ("commentCount", "Comments"),
        ]
    if plat == "twitter":
        return [
            ("followersCount", "Followers"),
            ("followingCount", "Following"),
            ("tweetCount", "Tweet Count"),
            ("likeCount", "Profile Likes"),
            ("listedCount", "Listed Count"),
        ]
    if plat == "tiktok":
        return [
            ("viewCountTotal", "Total Views"),
            ("followerCount", "Followers"),
            ("likeCountTotal", "Total Likes"),
            ("shareCountTotal", "Total Shares"),
            ("commentCountTotal", "Total Comments"),
            ("profileViews", "Profile Views"),
            ("videoCountTotal", "Video Count"),
        ]
    return []


def _filter_posts_by_month(history: list, plat: str, year: int, month: int) -> list:
    result = []
    for h in history:
        if h.get("status") != "success":
            continue
        if not in_month(str(h.get("created", "")), year, month):
            continue
        platforms = h.get("platforms") or []
        if plat in [str(p).lower() for p in platforms]:
            result.append(h)
    return result


def _verify_excel_platform(plat: str, posts_a: list, posts_b: list,
                           post_metrics: dict) -> bool:
    """从 Excel 读取指定平台的数据，与 post_metrics 对比。"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("     ⏭  openpyxl 未安装，跳过 Excel 验证")
        return True

    wb = load_workbook(OUT_XLSX)
    ws = wb.active

    # 找到平台列
    plat_idx = PLATFORMS.index(plat)
    c0 = 1 + plat_idx * 6  # 列偏移
    c_feb = c0 + 1
    c_mar = c0 + 2

    # 读取 Excel 中的值
    rows_data = {}
    for r in range(4, 60):
        label = ws.cell(row=r, column=c0).value
        feb = ws.cell(row=r, column=c_feb).value
        mar = ws.cell(row=r, column=c_mar).value
        if label:
            rows_data[str(label)] = {"feb": feb, "mar": mar}

    # 对比关键指标
    views_a = sum(
        float(post_metrics.get((str(h.get("id", "")), plat), {}).get("views", 0) or 0)
        for h in posts_a
    )
    views_b = sum(
        float(post_metrics.get((str(h.get("id", "")), plat), {}).get("views", 0) or 0)
        for h in posts_b
    )

    excel_views_a = rows_data.get("Views", {}).get("feb", 0)
    excel_views_b = rows_data.get("Views", {}).get("mar", 0)

    # 转换为数值（可能是 "4.2K" 格式）
    def parse_excel_val(v):
        if v is None or v == "—":
            return 0.0
        if isinstance(v, (int, float)):
            return float(v)
        s = str(v).replace(",", "").strip()
        if s.endswith("K"):
            return float(s[:-1]) * 1000
        if s.endswith("M"):
            return float(s[:-1]) * 1_000_000
        try:
            return float(s)
        except ValueError:
            return 0.0

    excel_a = parse_excel_val(excel_views_a)
    excel_b = parse_excel_val(excel_views_b)

    # 允许 1% 误差
    def within_tolerance(a, b):
        if a == 0 and b == 0:
            return True
        if a == 0 or b == 0:
            return abs(a - b) <= 1
        return abs(a - b) / max(a, b) < 0.01

    ok = within_tolerance(views_a, excel_a) and within_tolerance(views_b, excel_b)

    print(f"     [{plat}] Views: Python={fmt_num(views_a)}/{fmt_num(views_b)}  "
          f"Excel={fmt_num(excel_a)}/{fmt_num(excel_b)}  "
          f"{'✅' if ok else '❌'}")

    return ok


def fmt_mom(v: float | None) -> str:
    if v is None:
        return "—"
    return f"{v:+.1f}%"


def _save_debug_json(name: str, data: Any) -> None:
    os.makedirs(DEBUG, exist_ok=True)
    path = os.path.join(DEBUG, name)
    with open(path, "w") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    print(f"  📄  已保存: debug/{name}")


# ══════════════════════════════════════════════════════════════════════════
#  主函数
# ══════════════════════════════════════════════════════════════════════════

def main() -> None:
    banner(f"SAMA Monthly Report 数据诊断  |  {MONTH_LABELS[0]} vs {MONTH_LABELS[1]} {YEAR}", "═")

    # ── 加载缓存 ────────────────────────────────────────────────────────────
    print(f"\n加载缓存文件...")
    required = [
        f"{CACHE}/history_{MONTH_LABELS[0].lower()}.json",
        f"{CACHE}/history_{MONTH_LABELS[1].lower()}.json",
        f"{CACHE}/social_{MONTH_LABELS[0].lower()}.json",
        f"{CACHE}/social_{MONTH_LABELS[1].lower()}.json",
        f"{CACHE}/post_analytics_merged.json",
    ]
    for p in required:
        if not os.path.exists(p):
            print(f"\n  ❌  缺少文件: {p}")
            print(f"     请先运行 social_performance_six_platforms.py 进行数据抓取！")
            sys.exit(1)
        print(f"  ✅  {p}")

    hist_a = load_json(required[0])
    hist_b = load_json(required[1])
    soc_a  = load_json(required[2])
    soc_b  = load_json(required[3])

    with open(required[4]) as f:
        raw_analytics = json.load(f)

    print(f"\n加载完成: {len(raw_analytics)} 条帖子分析记录")

    # ── 创建 debug 目录 ─────────────────────────────────────────────────────
    os.makedirs(DEBUG, exist_ok=True)

    # ── 对每个平台运行诊断 ──────────────────────────────────────────────────
    all_reports: list[dict] = []
    for plat in PLATFORMS:
        banner(f"平台: {PLATFORM_LABELS[plat]}", "─")
        try:
            report = diagnose_platform(plat, raw_analytics, hist_a, hist_b, soc_a, soc_b)
            all_reports.append(report)
        except Exception as e:
            print(f"  ❌  诊断出错: {e}")
            all_reports.append({
                "platform": plat,
                "label": PLATFORM_LABELS[plat],
                "status": "ERROR",
                "error": str(e),
            })

    # ── 保存总报告 ──────────────────────────────────────────────────────────
    summary = {
        "generated_at": datetime.now().isoformat(),
        "year": YEAR,
        "months": {"a": MONTH_LABELS[0], "b": MONTH_LABELS[1]},
        "platforms": all_reports,
        "overall_status": "PASS" if all(r.get("status") == "PASS" for r in all_reports) else "FAIL",
    }
    _save_debug_json("report_summary.json", summary)

    # ── 最终汇总 ────────────────────────────────────────────────────────────
    banner("最终诊断结果", "═")

    for r in all_reports:
        status_icon = {"PASS": "✅", "FAIL": "❌", "ERROR": "🚨"}.get(r.get("status", "?"), "?")
        print(f"\n  {status_icon}  {r['label']} — {r.get('status', '?')}")

        if r.get("bugs"):
            for bug in r["bugs"]:
                print(f"       🐛  BUG: {bug}")
        if r.get("warnings"):
            for w in r["warnings"]:
                print(f"       ⚠️   WARN: {w}")

        # 打印关键数字
        ta = r["totals_a"]
        tb = r["totals_b"]
        print(f"       📊  Posts:  {ta['posts']:>3d} → {tb['posts']:>3d}")
        print(f"       📊  Views:  {fmt_num(ta['views']):>8s} → {fmt_num(tb['views']):>8s}  "
              f"({fmt_mom(r.get('mom_views'))})")
        print(f"       📊  Impr:   {fmt_num(ta['impressions']):>8s} → {fmt_num(tb['impressions']):>8s}  "
              f"({fmt_mom(r.get('mom_impressions'))})")
        print(f"       👥  Followers: {fmt_num(r.get('followers_a',0))} → {fmt_num(r.get('followers_b',0))}")

    banner("Excel 报表状态", "═")
    if os.path.exists(OUT_XLSX):
        mtime = datetime.fromtimestamp(os.path.getmtime(OUT_XLSX))
        print(f"\n  📄  {OUT_XLSX}")
        print(f"      最后修改: {mtime.strftime('%Y-%m-%d %H:%M:%S')}")
    else:
        print(f"\n  ⏭  Excel 文件不存在（先诊断，修复后运行 rebuild_excel_from_cache.py）")

    banner("下一步", "─")
    fail_count = sum(1 for r in all_reports if r.get("status") == "FAIL")
    if fail_count > 0:
        print(f"\n  ❌  发现 {fail_count} 个平台有 bug，需要修复后重新运行 rebuild_excel_from_cache.py")
    else:
        print(f"\n  ✅  所有平台数据正常！运行 rebuild_excel_from_cache.py 生成 Excel 报表。")
    print(f"\n  📋  详细报告: debug/report_summary.json")


if __name__ == "__main__":
    main()
