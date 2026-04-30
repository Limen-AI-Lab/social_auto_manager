#!/usr/bin/env python3
"""Rebuild Excel from cached data (no API calls)."""
import json, os, io, re
from collections import defaultdict
from datetime import datetime
import requests
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

YEAR = 2026
MONTH_A, MONTH_B = 2, 3
MONTH_LABELS = ("Feb", "Mar")
PLATFORMS = ["linkedin", "facebook", "instagram", "youtube", "twitter", "tiktok"]
PLATFORM_TITLE = {
    "linkedin": "LinkedIn", "facebook": "Facebook", "instagram": "Instagram",
    "youtube": "YouTube", "twitter": "X (Twitter)", "tiktok": "TikTok",
}
HEADER_HEX = {
    "linkedin": "9BC2E6", "facebook": "4472C4", "instagram": "E91E8C",
    "youtube": "C00000", "twitter": "1DA1F2", "tiktok": "000000",
}
CACHE = "scraped_report_cache"

def n(v):
    if v is None: return 0.0
    if isinstance(v, (int, float)) and not isinstance(v, bool): return float(v)
    return 0.0

def collect_posts(history, year, month):
    by = {p: [] for p in PLATFORMS}
    for h in history:
        if h.get("status") != "success": continue
        try:
            d = datetime.fromisoformat(str(h.get("created", "")).replace("Z", "+00:00"))
            if d.year != year or d.month != month: continue
        except Exception:
            continue
        for pl in h.get("platforms") or []:
            pl = str(pl).lower()
            if pl in by:
                by[pl].append(h)
    return by

def pct_change(p, c):
    if p == 0 and c == 0: return 0.0
    if p == 0: return None
    return (c - p) / p * 100.0

def fmt_mom(v):
    if v is None: return "—"
    return f"{v:.1f}%"

with open(f"{CACHE}/history_feb.json") as f:
    hist_a = json.load(f)
with open(f"{CACHE}/history_mar.json") as f:
    hist_b = json.load(f)
with open(f"{CACHE}/social_feb.json") as f:
    soc_a = json.load(f)
with open(f"{CACHE}/social_mar.json") as f:
    soc_b = json.load(f)
with open(f"{CACHE}/post_analytics_merged.json") as f:
    raw_analytics = json.load(f)

def extract(plat, body):
    out = {"impressions": 0.0, "reach": 0.0, "views": 0.0, "clicks": 0.0,
           "likes": 0.0, "comments": 0.0, "shares": 0.0, "reposts": 0.0}
    if body.get("status") == "error": return out
    pdata = body.get(plat)
    if isinstance(pdata, list) and pdata:
        pdata = pdata[0]
    if not isinstance(pdata, dict): return out
    a = pdata.get("analytics") or {}

    if plat == "facebook":
        out["impressions"] = n(a.get("impressionsUnique")) or n(a.get("impressionsOrganicUnique"))
        out["reach"] = out["impressions"]
        out["views"] = n(a.get("blueReelsPlayCount")) or out["impressions"]
        out["likes"] = n(a.get("likeCount"))
        out["comments"] = n(a.get("commentsCount"))
        out["shares"] = n(a.get("sharesCount")) or n(a.get("shareCount"))

    elif plat == "instagram":
        out["impressions"] = n(a.get("reachCount")) or n(a.get("impressionsCount")) or n(a.get("viewsCount"))
        out["reach"] = n(a.get("reachCount"))
        out["views"] = n(a.get("viewsCount"))
        out["likes"] = n(a.get("likeCount"))
        out["comments"] = n(a.get("commentsCount"))
        out["shares"] = n(a.get("sharesCount")) or n(a.get("shareCount"))
        out["reposts"] = out["shares"]

    elif plat == "linkedin":
        out["impressions"] = n(a.get("impressionCount"))
        out["reach"] = n(a.get("uniqueImpressionsCount"))
        out["views"] = n(a.get("videoViews"))
        out["clicks"] = n(a.get("clickCount"))
        out["likes"] = n(a.get("likeCount"))
        out["comments"] = n(a.get("commentCount"))
        out["shares"] = n(a.get("shareCount"))
        out["reposts"] = out["shares"]

    elif plat == "youtube":
        out["views"] = n(a.get("viewCount")) or n(a.get("views")) or n(a.get("videoViews"))
        out["likes"] = n(a.get("likeCount")) or n(a.get("likes"))
        out["comments"] = n(a.get("commentCount")) or n(a.get("comments"))
        out["shares"] = n(a.get("shareCount")) or n(a.get("shares"))
        out["reposts"] = out["shares"]
        out["impressions"] = out["views"]

    elif plat == "twitter":
        pub = a.get("publicMetrics") or {}
        out["impressions"] = n(pub.get("impressionCount")) or n(a.get("impressions"))
        out["likes"] = n(pub.get("likeCount")) or n(a.get("likeCount"))
        out["comments"] = n(pub.get("replyCount")) or n(a.get("replyCount"))
        out["reposts"] = n(pub.get("retweetCount")) or n(a.get("retweetCount"))
        out["shares"] = out["reposts"]
        out["clicks"] = n(a.get("urlLinkClicks")) or n(a.get("userProfileClicks"))

    elif plat == "tiktok":
        # TikTok API returns videoViews, NOT viewCount or views.
        # likeCount is the primary engagement metric.
        # commentsCount may be absent (API caps comment data).
        # shareCount is available for viral posts.
        # reach is available per post.
        # favorites is a separate metric (not in standard schema).
        out["views"]       = n(a.get("videoViews")) or n(a.get("viewCount")) or n(a.get("views"))
        out["likes"]       = n(a.get("likeCount")) or n(a.get("likes"))
        out["comments"]    = n(a.get("commentsCount")) or n(a.get("commentCount")) or n(a.get("comments"))
        out["shares"]      = n(a.get("shareCount")) or n(a.get("shares"))
        out["reposts"]     = out["shares"]
        out["reach"]       = n(a.get("reach")) or 0
        # TikTok has no separate "impressions" field — use views as best proxy
        out["impressions"] = out["views"]

    return out

pm = {}
for r in raw_analytics:
    pid, plat = r["post_id"], r["platform"]
    body = r.get("body", {})
    if body.get("status") == "success" and (pid, plat) not in pm:
        pm[(pid, plat)] = extract(plat, body)

def followers(plat, soc):
    block = soc.get(plat) or {}
    a = block.get("analytics") or {}
    if plat == "linkedin":   return n((a.get("followers") or {}).get("totalFollowerCount"))
    if plat == "facebook":   return n(a.get("followersCount"))
    if plat == "instagram":  return n(a.get("followersCount"))
    if plat == "youtube":    return n(a.get("subscriberCount"))
    if plat == "twitter":    return n(a.get("followersCount"))
    if plat == "tiktok":     return n(a.get("followerCount"))
    return 0.0

def soc_imp_proxy(plat, soc):
    block = soc.get(plat) or {}
    a = block.get("analytics") or {}
    if plat == "facebook":   return n(a.get("pagePostsImpressionsUnique")) or n(a.get("pagePostsImpressions"))
    if plat == "instagram":  return n(a.get("reachCount")) or n(a.get("viewsCount"))
    if plat == "linkedin":  return n(a.get("impressionCount"))
    if plat == "youtube":   return n(a.get("viewCount"))
    if plat == "tiktok":    return n(a.get("viewCountTotal"))
    return 0.0

def sum_month(posts, key, plat):
    s = 0.0
    for h in posts:
        pid = str(h.get("id", ""))
        s += float(pm.get((pid, plat), {}).get(key, 0) or 0)
    return s

by_a = collect_posts(hist_a, YEAR, MONTH_A)
by_b = collect_posts(hist_b, YEAR, MONTH_B)

# ── Verify data ──────────────────────────────────────────────────────────────
print("=== Per-platform post-level totals ===")
for plat in PLATFORMS:
    pa, pb = by_a.get(plat, []), by_b.get(plat, [])
    imp_a = sum_month(pa, "impressions", plat)
    imp_b = sum_month(pb, "impressions", plat)
    likes_a = sum_month(pa, "likes", plat)
    likes_b = sum_month(pb, "likes", plat)
    views_a = sum_month(pa, "views", plat)
    views_b = sum_month(pb, "views", plat)
    print(f"  {plat:12s}  Feb={len(pa)} posts imp={imp_a:,.0f} likes={likes_a:,.0f}  Mar={len(pb)} posts imp={imp_b:,.0f} likes={likes_b:,.0f}  views={views_b:,.0f}")

# ── Build Excel ──────────────────────────────────────────────────────────────

def thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

wb = Workbook()
ws = wb.active
ws.title = "Performance"
ws.sheet_view.showGridLines = False

def block_start(i): return 1 + i * 6

# Row 1 — title banner
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=block_start(6) + 4)
c = ws.cell(row=1, column=1,
            value=f"  Boolell Advisory Mauritius — Social Media Performance: {MONTH_LABELS[0]} vs {MONTH_LABELS[1]} {YEAR}")
c.font = Font(bold=True, color="FFFFFF", size=13)
c.fill = PatternFill("solid", fgColor="0D1B2A")
c.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 30

# Row 2 — note line
tot_a = sum(len(v) for v in by_a.values())
tot_b = sum(len(v) for v in by_b.values())
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=block_start(6) + 4)
c2 = ws.cell(row=2, column=1,
             value=f"  Live Ayrshare scrape | {MONTH_LABELS[0]}: {tot_a} posts ({sum(len(by_a[p]) for p in PLATFORMS)} total) | "
                   f"{MONTH_LABELS[1]}: {tot_b} posts | "
                   f"Post-level metrics summed across all posts in period; falls back to social-API period numbers if zero. "
                   f"MoM % Δ = ({MONTH_LABELS[1]}−{MONTH_LABELS[0]})/{MONTH_LABELS[0]}×100. "
                   f"Post of the Month chosen by highest impressions/views in {MONTH_LABELS[1]}.")
c2.font = Font(size=9, italic=True, color="555555")
c2.alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[2].height = 36

# Row 4 — platform name headers
for i, plat in enumerate(PLATFORMS):
    c0 = block_start(i)
    ws.merge_cells(start_row=4, start_column=c0, end_row=4, end_column=c0 + 3)
    h = ws.cell(row=4, column=c0, value=PLATFORM_TITLE[plat])
    h.font = Font(bold=True, color="FFFFFF", size=12)
    h.fill = PatternFill("solid", fgColor=HEADER_HEX[plat])
    h.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[4].height = 24

# Row 5 — metric col sub-headers
for i, plat in enumerate(PLATFORMS):
    c0 = block_start(i)
    for ci, label in enumerate(["Metric", MONTH_LABELS[0], MONTH_LABELS[1], "MoM % Δ"]):
        ch = ws.cell(row=5, column=c0 + ci, value=label)
        ch.font = Font(bold=True)
        ch.border = thin()
        ch.alignment = Alignment(horizontal="center")
ws.row_dimensions[5].height = 22

def metric_specs(plat):
    if plat == "linkedin":
        return [("Number of Posts","posts"), ("Followers","followers"), ("Follower Growth %","fg"),
                ("Impressions","imp"), ("Reach","reach"), ("Clicks","clicks"),
                ("Reactions","likes"), ("Reposts","shares"), ("Engagement %","eng_pct")]
    if plat == "facebook":
        return [("Number of Posts","posts"), ("Followers","followers"), ("Follower Growth %","fg"),
                ("Impressions","imp"), ("Reach","reach"), ("Clicks","clicks"),
                ("Reactions","likes"), ("Shares","shares")]
    if plat == "instagram":
        return [("Number of Posts","posts"), ("Followers","followers"), ("Follower Growth %","fg"),
                ("Impressions","imp"), ("Reach","reach"), ("Views","views"),
                ("Reactions","likes"), ("Shares","shares")]
    if plat == "youtube":
        return [("Number of Posts","posts"), ("Followers","followers"), ("Follower Growth %","fg"),
                ("Views","views"), ("Likes","likes"), ("Comments","comments"),
                ("Shares","shares"), ("Engagement","eng_yt")]
    if plat == "twitter":
        return [("Number of Posts","posts"), ("Followers","followers"), ("Follower Growth %","fg"),
                ("Impressions","imp"), ("Clicks","clicks"), ("Likes","likes"), ("Reposts","shares")]
    # tiktok — views is primary; reach, comments, shares are secondary
    return [("Number of Posts","posts"), ("Followers","followers"), ("Follower Growth %","fg"),
            ("Views","views"), ("Reach","reach"), ("Likes","likes"),
            ("Comments","comments"), ("Shares","shares")]

DATA_ROW = 6

for i, plat in enumerate(PLATFORMS):
    c0 = block_start(i)
    pa, pb = by_a.get(plat, []), by_b.get(plat, [])
    na, nb = len(pa), len(pb)

    fa, fb = followers(plat, soc_a), followers(plat, soc_b)
    fg_pct = pct_change(fa, fb)

    imp_a  = sum_month(pa, "impressions", plat) or soc_imp_proxy(plat, soc_a)
    imp_b  = sum_month(pb, "impressions", plat) or soc_imp_proxy(plat, soc_b)
    reach_a = sum_month(pa, "reach", plat)
    reach_b = sum_month(pb, "reach", plat)
    views_a = sum_month(pa, "views", plat)
    views_b = sum_month(pb, "views", plat)
    clicks_a = sum_month(pa, "clicks", plat)
    clicks_b = sum_month(pb, "clicks", plat)
    likes_a  = sum_month(pa, "likes", plat)
    likes_b  = sum_month(pb, "likes", plat)
    comms_a  = sum_month(pa, "comments", plat)
    comms_b  = sum_month(pb, "comments", plat)
    shares_a = sum_month(pa, "shares", plat)
    shares_b = sum_month(pb, "shares", plat)
    eng_a    = 100.0 * (likes_a + comms_a + shares_a) / imp_a if imp_a > 0 else 0.0
    eng_b    = 100.0 * (likes_b + comms_b + shares_b) / imp_b if imp_b > 0 else 0.0
    eng_yt_a = likes_a + comms_a + shares_a
    eng_yt_b = likes_b + comms_b + shares_b

    V = {
        "posts": (na, nb), "followers": (fa, fb), "fg": (None, fg_pct),
        "imp":  (imp_a,  imp_b),
        "reach":(reach_a, reach_b),
        "views":(views_a, views_b),
        "clicks":(clicks_a, clicks_b),
        "likes":(likes_a, likes_b),
        "comments":(comms_a, comms_b),
        "shares":(shares_a, shares_b),
        "eng_pct":(eng_a, eng_b),
        "eng_yt":(eng_yt_a, eng_yt_b),
    }

    row = DATA_ROW
    for label, key in metric_specs(plat):
        va, vb = V[key]
        if key == "fg":
            mom_cell = ws.cell(row=row, column=c0 + 3, value="—")
        else:
            mom = pct_change(float(va), float(vb))
            mom_cell = ws.cell(row=row, column=c0 + 3, value=fmt_mom(mom))
            if mom is not None and mom > 0:
                mom_cell.fill = PatternFill("solid", fgColor="C6EFCE")
            elif mom is not None and mom < 0:
                mom_cell.fill = PatternFill("solid", fgColor="FFC7CE")

        ws.cell(row=row, column=c0, value=label).border = thin()
        ca = ws.cell(row=row, column=c0 + 1, value=va)
        cb = ws.cell(row=row, column=c0 + 2, value=vb)
        for cell in (ca, cb, mom_cell):
            cell.border = thin()
            cell.alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=c0).alignment = Alignment(horizontal="left")
        row += 1

    # ── Post of the Month ────────────────────────────────────────────────────
    row += 1
    banner = ws.cell(row=row, column=c0, value="POST OF THE MONTH")
    banner.font = Font(bold=True, color="FFFFFF", size=10)
    banner.fill = PatternFill("solid", fgColor=HEADER_HEX[plat])
    ws.merge_cells(start_row=row, start_column=c0 + 1, end_row=row, end_column=c0 + 3)
    row += 1

    # Pick top post in Mar (month B) by score.
    # TikTok: score = views*0.4 + engagements*0.4 + reach*0.2 (fallback: views only).
    # Other platforms: score = impressions or views.
    top_h = None
    top_score = -1.0
    for h in pb:
        pid = str(h.get("id", ""))
        met = pm.get((pid, plat), {})
        if plat == "tiktok":
            views = met.get("views", 0) or 0
            engagements = (met.get("likes", 0) or 0) + (met.get("comments", 0) or 0) + (met.get("shares", 0) or 0)
            reach = met.get("reach", 0) or 0
            score = views * 0.4 + engagements * 0.4 + reach * 0.2 if views > 0 else -1.0
        else:
            score = met.get("impressions", 0) or met.get("views", 0)
        if score > top_score:
            top_score = score
            top_h = h

    # Post link
    url = ""
    if top_h:
        for entry in top_h.get("postIds") or []:
            if str(entry.get("platform", "")).lower() == plat:
                url = str(entry.get("postUrl", ""))
                break
    link_lbl = ws.cell(row=row, column=c0, value="Link")
    link_lbl.font = Font(bold=True); link_lbl.border = thin()
    ws.merge_cells(start_row=row, start_column=c0 + 1, end_row=row, end_column=c0 + 3)
    link_cell = ws.cell(row=row, column=c0 + 1, value="Open post" if url else "(no link)")
    link_cell.border = thin()
    if url:
        link_cell.hyperlink = url
        link_cell.font = Font(color="0563C1", underline="single")
    row += 1

    # Topic
    topic = ""
    if top_h:
        text = (top_h.get("post", "") or "").strip()
        topic = text.split("\n")[0][:100]
    topic_lbl = ws.cell(row=row, column=c0, value="Topic")
    topic_lbl.font = Font(bold=True); topic_lbl.border = thin()
    ws.merge_cells(start_row=row, start_column=c0 + 1, end_row=row, end_column=c0 + 3)
    tc = ws.cell(row=row, column=c0 + 1, value=topic)
    tc.border = thin(); tc.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[row].height = 28; row += 1

    # Narrative
    narrative = ""
    if top_h:
        narrative = (top_h.get("post", "") or "")[:2000]
    narr_lbl = ws.cell(row=row, column=c0, value="Narrative")
    narr_lbl.font = Font(bold=True); narr_lbl.border = thin()
    # merge to cover 2 rows
    ws.merge_cells(start_row=row, start_column=c0 + 1, end_row=row + 1, end_column=c0 + 3)
    nc = ws.cell(row=row, column=c0 + 1, value=narrative)
    nc.border = thin(); nc.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = 14; ws.row_dimensions[row + 1].height = 14
    row += 2

    # Post-level metrics
    met = (pm.get((str(top_h.get("id", "")), plat), {}) if top_h else {})
    if plat == "linkedin":
        rows_pm = [("Impressions", met.get("impressions", 0)), ("Clicks", met.get("clicks", 0))]
        ctr = f"{round(100.0 * met.get('clicks', 0) / met['impressions'], 2)}%" if met.get("impressions") else "0%"
        rows_pm.append(("CTR", ctr))
    elif plat == "facebook":
        rows_pm = [("Impressions", met.get("impressions", 0)),
                   ("Views/Reactions", met.get("views", 0) + met.get("likes", 0))]
    elif plat == "instagram":
        rows_pm = [("Impressions", met.get("impressions", 0)), ("Reach", met.get("reach", 0)),
                   ("Views", met.get("views", 0))]
    elif plat == "youtube":
        rows_pm = [("Views", met.get("views", 0)), ("Likes", met.get("likes", 0)),
                   ("Comments", met.get("comments", 0)), ("Shares", met.get("shares", 0))]
    elif plat == "twitter":
        rows_pm = [("Impressions", met.get("impressions", 0)), ("Likes", met.get("likes", 0)),
                   ("Reposts", met.get("reposts", 0))]
    else:
        rows_pm = [("Views", met.get("views", 0)), ("Likes", met.get("likes", 0)),
                   ("Comments", met.get("comments", 0)), ("Shares", met.get("shares", 0))]

    for lbl, val in rows_pm:
        ml = ws.cell(row=row, column=c0, value=lbl)
        ml.font = Font(bold=True); ml.border = thin()
        ws.merge_cells(start_row=row, start_column=c0 + 1, end_row=row, end_column=c0 + 3)
        mv = ws.cell(row=row, column=c0 + 1, value=val)
        mv.border = thin(); row += 1

    # Thumbnail
    img_ok = False
    if top_h and top_h.get("mediaUrls"):
        img_url = str(top_h["mediaUrls"][0])
        if re.search(r"\.(jpe?g|png|gif|webp)(\?|$)", img_url, re.I):
            try:
                ir = requests.get(img_url, timeout=25)
                if ir.ok and ir.content:
                    bio = io.BytesIO(ir.content)
                    img = XLImage(bio)
                    img.width = min(img.width, 220); img.height = min(img.height, 150)
                    ws.add_image(img, f"{get_column_letter(c0 + 1)}{row}")
                    ws.row_dimensions[row].height = 120; img_ok = True
            except Exception:
                pass
    if not img_ok:
        il = ws.cell(row=row, column=c0, value="Image")
        il.font = Font(bold=True); il.border = thin()
        ws.merge_cells(start_row=row, start_column=c0 + 1, end_row=row, end_column=c0 + 3)
        iv = ws.cell(row=row, column=c0 + 1, value="(preview unavailable — video or blocked URL)")
        iv.border = thin(); ws.row_dimensions[row].height = 18

    # Column widths
    for cc in range(c0, c0 + 4):
        ws.column_dimensions[get_column_letter(cc)].width = 14 if cc > c0 else 22

wb.save("Social_Media_Performance_Feb_Mar_2026_6Platforms.xlsx")
print("\nWrote: Social_Media_Performance_Feb_Mar_2026_6Platforms.xlsx")
