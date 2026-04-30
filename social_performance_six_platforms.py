#!/usr/bin/env python3
"""
Fresh scrape from Ayrshare + Excel in the 6-column Social Media Performance layout
(Jan | Feb | MoM % Δ per platform + Post of the Month per platform).

Env: AYRSHARE_API_KEY, AYRSHARE_PROFILE_KEY (optional; falls back to same keys as other scripts).
X_API_KEY, X_API_SECRET: X/Twitter BYO API credentials (required since March 31, 2026).
https://www.ayrshare.com/docs/dashboard/connect-social-accounts/x-twitter-byo-keys

Note: If /api/history returns no rows for January, post counts and post-level sums for Jan stay 0;
      social snapshots for the Jan date range are still requested. A notes row is added on the sheet.
"""

from __future__ import annotations

import io
import json
import os
import re
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from typing import Any

import requests
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
# ── Config ───────────────────────────────────────────────────────────────────

API_KEY = os.environ.get("AYRSHARE_API_KEY", "788C1419-FB69401D-96E97381-FD75B910")
PROFILE_KEY = os.environ.get("AYRSHARE_PROFILE_KEY", "805E0090-79E243F0-8157B77C-42D928FB")
# X/Twitter BYO API credentials (required since March 31, 2026)
X_API_KEY = os.environ.get("X_API_KEY", "")
X_API_SECRET = os.environ.get("X_API_SECRET", "")
BASE = "https://api.ayrshare.com"

YEAR = 2026
MONTH_A = 2   # Feb (label col 1)
MONTH_B = 3   # Mar (label col 2)
MONTH_LABELS = ("Feb", "Mar")

PLATFORMS = ["linkedin", "facebook", "instagram", "youtube", "twitter", "tiktok"]
PLATFORM_TITLE = {
    "linkedin": "LinkedIn",
    "facebook": "Facebook",
    "instagram": "Instagram",
    "youtube": "YouTube",
    "twitter": "X (Twitter)",
    "tiktok": "TikTok",
}
HEADER_HEX = {
    "linkedin": "9BC2E6",
    "facebook": "4472C4",
    "instagram": "E91E8C",
    "youtube": "C00000",
    "twitter": "1DA1F2",
    "tiktok": "000000",
}

OUT_JSON_DIR = "scraped_report_cache"
OUT_XLSX = "Social_Media_Performance_Feb_Mar_2026_6Platforms.xlsx"
MAX_WORKERS = 10
POST_TIMEOUT = 45
MAX_RETRIES = 3

# ── Dates ────────────────────────────────────────────────────────────────────


def month_range(year: int, month: int) -> tuple[str, str]:
    if month == 12:
        nxt = f"{year + 1}-01-01"
    else:
        nxt = f"{year}-{month + 1:02d}-01"
    start = f"{year}-{month:02d}-01"
    # end = last day
    if month in (1, 3, 5, 7, 8, 10, 12):
        end = f"{year}-{month:02d}-31"
    elif month in (4, 6, 9, 11):
        end = f"{year}-{month:02d}-30"
    else:
        end = f"{year}-02-29" if year % 4 == 0 and (year % 100 != 0 or year % 400 == 0) else f"{year}-02-28"
    return start, end


def in_month(iso_ts: str, year: int, month: int) -> bool:
    if not iso_ts:
        return False
    try:
        d = datetime.fromisoformat(iso_ts.replace("Z", "+00:00"))
    except ValueError:
        return False
    return d.year == year and d.month == month


def n(v: Any) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    return 0.0


# ── HTTP ───────────────────────────────────────────────────────────────────────


def _build_headers() -> dict[str, str]:
    """Build request headers with optional X/Twitter BYO credentials."""
    headers = {
        "Authorization": f"Bearer {API_KEY}",
        "Content-Type": "application/json",
    }
    if X_API_KEY:
        headers["X-Twitter-OAuth1-Api-Key"] = X_API_KEY
    if X_API_SECRET:
        headers["X-Twitter-OAuth1-Api-Secret"] = X_API_SECRET
    return headers


def _session() -> requests.Session:
    s = requests.Session()
    s.headers.update(_build_headers())
    return s


def fetch_history(sess: requests.Session, start: str, end: str) -> list[dict[str, Any]]:
    r = sess.get(
        f"{BASE}/api/history",
        params={"profileKey": PROFILE_KEY, "limit": "500", "startDate": start, "endDate": end},
        timeout=120,
    )
    data = r.json()
    h = data.get("history")
    if isinstance(h, list):
        return h
    return []


def fetch_social(sess: requests.Session, start: str, end: str) -> dict[str, Any]:
    last_err = ""
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            r = sess.post(
                f"{BASE}/api/analytics/social",
                json={
                    "profileKey": PROFILE_KEY,
                    "platforms": PLATFORMS,
                    "startDate": start,
                    "endDate": end,
                },
                timeout=300,
            )
            r.raise_for_status()
            return r.json()
        except Exception as e:
            last_err = str(e)
            print(f"    ⚠ fetch_social attempt {attempt}/{MAX_RETRIES} failed: {last_err[:120]}")
            if attempt < MAX_RETRIES:
                time.sleep(3 * attempt)
    raise RuntimeError(f"fetch_social failed after {MAX_RETRIES} attempts: {last_err}")


def fetch_post_analytics(sess: requests.Session, post_id: str, platform: str) -> dict[str, Any]:
    r = sess.post(
        f"{BASE}/api/analytics/post",
        json={"id": post_id, "profileKey": PROFILE_KEY, "platforms": [platform]},
        timeout=POST_TIMEOUT,
    )
    return r.json()


# ── Extract social (account-level for period) ────────────────────────────────


def followers_from_social(plat: str, social_root: dict[str, Any]) -> float:
    block = social_root.get(plat) or {}
    a = block.get("analytics") or {}
    if plat == "linkedin":
        return n((a.get("followers") or {}).get("totalFollowerCount"))
    if plat == "facebook":
        return n(a.get("followersCount"))
    if plat == "instagram":
        return n(a.get("followersCount"))
    if plat == "youtube":
        return n(a.get("subscriberCount"))
    if plat == "twitter":
        return n(a.get("followersCount"))
    if plat == "tiktok":
        return n(a.get("followerCount"))
    return 0.0


def social_impressions_proxy(plat: str, social_root: dict[str, Any]) -> float:
    """Use API account/period fields where available (varies by platform)."""
    block = social_root.get(plat) or {}
    a = block.get("analytics") or {}
    if plat == "facebook":
        return n(a.get("pagePostsImpressionsUnique")) or n(a.get("pagePostsImpressions"))
    if plat == "instagram":
        return n(a.get("reachCount")) or n(a.get("viewsCount"))
    if plat == "linkedin":
        return n(a.get("impressionCount"))
    if plat == "youtube":
        return n(a.get("viewCount"))
    if plat == "twitter":
        return n(a.get("tweetCount"))
    if plat == "tiktok":
        # TikTok social analytics returns viewCountTotal for period views
        return n(a.get("viewCountTotal"))
    return 0.0


# ── Extract post-level ───────────────────────────────────────────────────────


def extract_post_metrics(platform: str, body: dict[str, Any]) -> dict[str, float]:
    out = {
        "impressions": 0.0,
        "reach": 0.0,
        "views": 0.0,
        "clicks": 0.0,
        "likes": 0.0,
        "comments": 0.0,
        "shares": 0.0,
        "reposts": 0.0,
        "engagement_rate": 0.0,
    }
    if body.get("status") == "error":
        return out
    plat = platform.lower()
    pdata = body.get(plat)
    if isinstance(pdata, list) and pdata:
        pdata = pdata[0]
    if not isinstance(pdata, dict):
        return out
    a = pdata.get("analytics") or {}
    if plat == "facebook":
        out["impressions"] = n(a.get("impressionsUnique")) or n(a.get("impressionsOrganicUnique"))
        out["reach"] = out["impressions"]
        out["views"] = n(a.get("blueReelsPlayCount")) or out["impressions"]
        out["likes"] = n(a.get("likeCount"))
        out["comments"] = n(a.get("commentsCount"))
        out["shares"] = n(a.get("sharesCount")) or n(a.get("shareCount"))
    elif plat == "instagram":
        out["impressions"] = n(a.get("reachCount")) or n(a.get("impressionsCount")) or n(a.get("viewsCount"))
        out["reach"] = n(a.get("reachCount"))
        out["views"] = n(a.get("viewsCount"))
        out["likes"] = n(a.get("likeCount")) or n(a.get("likes"))
        out["comments"] = n(a.get("commentsCount")) or n(a.get("comments"))
        out["shares"] = n(a.get("sharesCount")) or n(a.get("shareCount"))
        out["reposts"] = out["shares"]
    elif plat == "linkedin":
        out["impressions"] = n(a.get("impressionCount")) or n(a.get("uniqueImpressionsCount"))
        out["reach"] = n(a.get("uniqueImpressionsCount"))
        out["views"] = n(a.get("videoViews"))
        out["clicks"] = n(a.get("clickCount"))
        out["likes"] = n(a.get("likeCount"))
        out["comments"] = n(a.get("commentCount"))
        out["shares"] = n(a.get("shareCount"))
        out["reposts"] = out["shares"]
        eng = a.get("engagement")
        if isinstance(eng, (int, float)):
            out["engagement_rate"] = float(eng) * 100 if eng <= 1 else float(eng)
        elif out["impressions"] > 0:
            reacts = a.get("reactions") or {}
            if isinstance(reacts, dict):
                rsum = sum(n(v) for v in reacts.values())
            else:
                rsum = 0
            eng_raw = out["likes"] + out["comments"] + out["shares"] + rsum
            out["engagement_rate"] = 100.0 * eng_raw / out["impressions"]
    elif plat == "youtube":
        # viewCount / likeCount / comments / shares live at top level of analytics block
        out["views"] = n(a.get("viewCount")) or n(a.get("views")) or n(a.get("videoViews"))
        out["likes"] = n(a.get("likeCount")) or n(a.get("likes")) or n(a.get("reactions", {}).get("like", 0))
        out["comments"] = n(a.get("commentCount")) or n(a.get("comments"))
        out["shares"] = n(a.get("shareCount")) or n(a.get("shares"))
        out["reposts"] = out["shares"]
        out["impressions"] = out["views"]  # YouTube: views = impressions
    elif plat == "twitter":
        # Metrics live under publicMetrics
        pub = a.get("publicMetrics") or {}
        out["impressions"] = n(pub.get("impressionCount")) or n(a.get("impressions"))
        out["likes"] = n(pub.get("likeCount")) or n(a.get("likeCount")) or n(a.get("favoriteCount"))
        out["comments"] = n(pub.get("replyCount")) or n(a.get("replyCount")) or n(a.get("replies"))
        out["reposts"] = n(pub.get("retweetCount")) or n(a.get("retweetCount")) or n(a.get("retweets", 0))
        out["shares"] = out["reposts"]
        out["clicks"] = n(a.get("urlLinkClicks")) or n(a.get("userProfileClicks"))
    elif plat == "tiktok":
        # TikTok API returns videoViews, NOT viewCount or views.
        # likes is the primary engagement metric.
        # commentsCount may be absent or capped.
        # shareCount is available for viral posts.
        # reach is available per post.
        # favorites is available but not in standard output schema.
        out["views"]       = n(a.get("videoViews")) or n(a.get("viewCount")) or n(a.get("views"))
        out["likes"]       = n(a.get("likeCount")) or n(a.get("likes"))
        out["comments"]    = n(a.get("commentsCount")) or n(a.get("commentCount")) or n(a.get("comments"))
        out["shares"]      = n(a.get("shareCount")) or n(a.get("shares"))
        out["reposts"]     = out["shares"]
        out["reach"]       = n(a.get("reach")) or 0
        # TikTok has no separate "impressions" field — use views as best proxy
        out["impressions"] = out["views"]
    return out


def post_url_from_history(hpost: dict[str, Any], platform: str) -> str:
    for e in hpost.get("postIds") or []:
        if str(e.get("platform", "")).lower() == platform:
            return str(e.get("postUrl") or "")
    return ""


def topic_from_text(text: str, max_len: int = 90) -> str:
    text = (text or "").strip()
    if not text:
        return ""
    line = text.split("\n")[0].strip()
    if len(line) > max_len:
        return line[: max_len - 1] + "…"
    return line


# ── Aggregation ──────────────────────────────────────────────────────────────


def collect_posts_for_month(
    history: list[dict[str, Any]], year: int, month: int
) -> dict[str, list[dict[str, Any]]]:
    by_plat: dict[str, list[dict[str, Any]]] = {p: [] for p in PLATFORMS}
    for h in history:
        if h.get("status") != "success":
            continue
        if not in_month(str(h.get("created", "")), year, month):
            continue
        for pl in h.get("platforms") or []:
            pl = str(pl).lower()
            if pl in by_plat:
                by_plat[pl].append(h)
    return by_plat


def pct_change(prev: float, curr: float) -> float | None:
    if prev == 0 and curr == 0:
        return 0.0
    if prev == 0:
        return None
    return (curr - prev) / prev * 100.0


def fmt_mom(v: float | None) -> str:
    if v is None:
        return "—"
    return f"{v:.1f}%"


# ── Excel ─────────────────────────────────────────────────────────────────────


def thin() -> Border:
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def build_excel(
    month_a_posts: dict[str, list[dict]],
    month_b_posts: dict[str, list[dict]],
    social_a: dict[str, Any],
    social_b: dict[str, Any],
    post_metrics: dict[tuple[str, str], dict[str, float]],
    notes: str,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Performance"

    def block_start(idx: int) -> int:
        return 1 + idx * 6  # 5 cols + 1 gap

    ws.row_dimensions[1].height = 28
    note_row = 2
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=block_start(6) + 4)
    cell = ws.cell(row=note_row, column=1, value=notes)
    cell.font = Font(size=9, italic=True, color="555555")
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[note_row].height = 36

    header_row = 4
    for i, plat in enumerate(PLATFORMS):
        c0 = block_start(i)
        c1 = c0 + 3
        ws.merge_cells(
            start_row=header_row,
            start_column=c0,
            end_row=header_row,
            end_column=c1,
        )
        h = ws.cell(row=header_row, column=c0, value=PLATFORM_TITLE[plat])
        h.font = Font(bold=True, color="FFFFFF", size=12)
        h.fill = PatternFill("solid", fgColor=HEADER_HEX[plat])
        h.alignment = Alignment(horizontal="center", vertical="center")

    table_header_row = header_row + 1
    for i, plat in enumerate(PLATFORMS):
        c0 = block_start(i)
        ws.cell(row=table_header_row, column=c0, value="Metric").font = Font(bold=True)
        ws.cell(row=table_header_row, column=c0 + 1, value=MONTH_LABELS[0]).font = Font(bold=True)
        ws.cell(row=table_header_row, column=c0 + 2, value=MONTH_LABELS[1]).font = Font(bold=True)
        ws.cell(row=table_header_row, column=c0 + 3, value="MoM % Δ").font = Font(bold=True)
        for cc in range(c0, c0 + 4):
            ws.cell(row=table_header_row, column=cc).border = thin()
            ws.cell(row=table_header_row, column=cc).alignment = Alignment(horizontal="center")

    # Metric row definitions per platform (labels + keys for computed dict)
    def metric_specs(plat: str) -> list[tuple[str, str]]:
        common_posts = ("Number of Posts", "posts")
        common_fol = ("Followers", "followers")
        common_fg = ("Follower Growth %", "follower_growth_pct")
        if plat == "linkedin":
            return [
                common_posts,
                common_fol,
                common_fg,
                ("Impressions", "impressions"),
                ("Engagement %", "engagement_pct"),
                ("Clicks", "clicks"),
                ("Reactions", "reactions"),
                ("Reposts", "reposts"),
            ]
        if plat == "facebook":
            return [
                common_posts,
                common_fol,
                common_fg,
                ("Impressions", "impressions"),
                ("Reach", "reach"),
                ("Clicks", "clicks"),
                ("Reactions", "reactions"),
                ("Shares", "shares"),
            ]
        if plat == "instagram":
            return [
                common_posts,
                common_fol,
                common_fg,
                ("Total Impressions", "impressions"),
                ("Reach", "reach"),
                ("Clicks", "clicks"),
                ("Reactions", "reactions"),
                ("Shares", "shares"),
            ]
        if plat == "youtube":
            return [
                common_posts,
                common_fol,
                common_fg,
                ("Impressions", "impressions"),
                ("Views", "views"),
                ("Engagements", "engagements"),
            ]
        if plat == "twitter":
            return [
                common_posts,
                common_fol,
                common_fg,
                ("Impressions", "impressions"),
                ("Engagement %", "engagement_pct"),
                ("Clicks", "clicks"),
                ("Likes", "likes"),
                ("Reposts", "reposts"),
            ]
        # tiktok — views is primary; reach, comments, shares are secondary
        return [
            common_posts,
            common_fol,
            common_fg,
            ("Views", "views"),
            ("Reach", "reach"),
            ("Likes", "likes"),
            ("Comments", "comments"),
            ("Shares", "shares"),
        ]

    post_section_start_row = table_header_row + 1
    max_metric_rows = max(len(metric_specs(p)) for p in PLATFORMS)

    for i, plat in enumerate(PLATFORMS):
        c0 = block_start(i)
        pa = month_a_posts.get(plat, [])
        pb = month_b_posts.get(plat, [])
        posts_a = len(pa)
        posts_b = len(pb)

        fa = followers_from_social(plat, social_a)
        fb = followers_from_social(plat, social_b)
        fg_feb_pct = pct_change(fa, fb)

        # Sum post-level metrics for each month
        def sum_month(posts: list[dict], mkey: str) -> float:
            s = 0.0
            for h in posts:
                pid = h.get("id")
                if not pid:
                    continue
                met = post_metrics.get((str(pid), plat), {})
                s += float(met.get(mkey, 0) or 0)
            return s

        imp_a = sum_month(pa, "impressions") or social_impressions_proxy(plat, social_a)
        imp_b = sum_month(pb, "impressions") or social_impressions_proxy(plat, social_b)
        # If still zero on summed posts, fall back to social proxy for that month
        if posts_a and sum_month(pa, "impressions") == 0:
            imp_a = social_impressions_proxy(plat, social_a)
        if posts_b and sum_month(pb, "impressions") == 0:
            imp_b = social_impressions_proxy(plat, social_b)

        reach_a = sum_month(pa, "reach")
        reach_b = sum_month(pb, "reach")
        views_a = sum_month(pa, "views")
        views_b = sum_month(pb, "views")
        clicks_a = sum_month(pa, "clicks")
        clicks_b = sum_month(pb, "clicks")
        likes_a = sum_month(pa, "likes")
        likes_b = sum_month(pb, "likes")
        comments_a = sum_month(pa, "comments")
        comments_b = sum_month(pb, "comments")
        shares_a = sum_month(pa, "shares")
        shares_b = sum_month(pb, "shares")
        reposts_a = sum_month(pa, "reposts")
        reposts_b = sum_month(pb, "reposts")

        eng_a = (
            100.0 * (likes_a + comments_a + shares_a) / imp_a if imp_a > 0 else sum_month(pa, "engagement_rate")
        )
        eng_b = (
            100.0 * (likes_b + comments_b + shares_b) / imp_b if imp_b > 0 else sum_month(pb, "engagement_rate")
        )
        eng_a = eng_a if isinstance(eng_a, (int, float)) else 0
        eng_b = eng_b if isinstance(eng_b, (int, float)) else 0

        react_a = likes_a
        react_b = likes_b

        eng_yt_a = likes_a + comments_a + shares_a
        eng_yt_b = likes_b + comments_b + shares_b

        values = {
            "posts": (posts_a, posts_b),
            "followers": (fa, fb),
            "follower_growth_pct": (None, fg_feb_pct),
            "impressions": (imp_a, imp_b),
            "reach": (reach_a, reach_b),
            "views": (views_a, views_b),
            "clicks": (clicks_a, clicks_b),
            "likes": (likes_a, likes_b),
            "comments": (comments_a, comments_b),
            "shares": (shares_a, shares_b),
            "reposts": (reposts_a, reposts_b),
            "reactions": (react_a, react_b),
            "engagement_pct": (eng_a, eng_b),
            "engagements": (eng_yt_a, eng_yt_b),
        }

        row = post_section_start_row
        for label, key in metric_specs(plat):
            va, vb = values[key]
            if key == "follower_growth_pct":
                mom = None
                cell_mom = ws.cell(row=row, column=c0 + 3, value="—")
                disp_a = "—"
                disp_b = fmt_mom(vb) if vb is not None else "—"
            else:
                mom = pct_change(float(va), float(vb))
                cell_mom = ws.cell(row=row, column=c0 + 3, value=fmt_mom(mom))
                disp_a, disp_b = va, vb

            ws.cell(row=row, column=c0, value=label).border = thin()
            ca = ws.cell(row=row, column=c0 + 1, value=disp_a)
            cb = ws.cell(row=row, column=c0 + 2, value=disp_b)
            for c in (ca, cb, cell_mom):
                c.border = thin()
                c.alignment = Alignment(horizontal="right")
            ws.cell(row=row, column=c0).alignment = Alignment(horizontal="left")

            if mom is not None and key != "follower_growth_pct":
                if mom > 0:
                    cell_mom.fill = PatternFill("solid", fgColor="C6EFCE")
                elif mom < 0:
                    cell_mom.fill = PatternFill("solid", fgColor="FFC7CE")
            row += 1

        # Post of the month (best in Feb = month B)
        top_h = None
        top_score = -1.0
        for h in pb:
            pid = str(h.get("id", ""))
            met = post_metrics.get((pid, plat), {})
            score = met.get("impressions", 0) or met.get("views", 0)
            if score > top_score:
                top_score = score
                top_h = h

        post_block_row = post_section_start_row + max_metric_rows + 1
        ws.cell(row=post_block_row, column=c0, value="Link to Post of the Month").font = Font(
            bold=True, color="0563C1"
        )
        link_url = post_url_from_history(top_h, plat) if top_h else ""
        link_cell = ws.cell(row=post_block_row, column=c0 + 1, value="Open post" if link_url else "")
        if link_url:
            link_cell.hyperlink = link_url
            link_cell.font = Font(color="0563C1", underline="single")
        ws.merge_cells(
            start_row=post_block_row, start_column=c0 + 1, end_row=post_block_row, end_column=c0 + 3
        )

        pr = post_block_row + 1
        narrative = (top_h or {}).get("post", "") if top_h else ""
        topic = topic_from_text(narrative) if top_h else ""

        ws.cell(row=pr, column=c0, value="Topic").font = Font(bold=True)
        ws.merge_cells(start_row=pr, start_column=c0 + 1, end_row=pr, end_column=c0 + 3)
        ws.cell(row=pr, column=c0 + 1, value=topic).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[pr].height = 28

        pr += 1
        ws.cell(row=pr, column=c0, value="Narrative").font = Font(bold=True)
        ws.merge_cells(start_row=pr, start_column=c0 + 1, end_row=pr + 2, end_column=c0 + 3)
        narr_cell = ws.cell(row=pr, column=c0 + 1, value=narrative[:2000] if narrative else "")
        narr_cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[pr].height = 14
        ws.row_dimensions[pr + 1].height = 14
        ws.row_dimensions[pr + 2].height = 14

        pr += 3
        # Post-level metrics (Feb top post)
        met = (
            post_metrics.get((str(top_h.get("id")), plat), {})
            if top_h
            else {k: 0 for k in ["impressions", "views", "clicks", "reach", "likes", "shares", "comments"]}
        )
        if plat == "linkedin":
            rows_pm = [
                ("Impressions", met.get("impressions", 0)),
                ("Views", met.get("views", 0)),
                ("Clicks", met.get("clicks", 0)),
                (
                    "CTR",
                    f"{100.0 * met.get('clicks', 0) / met['impressions']:.2f}%"
                    if met.get("impressions")
                    else "0%",
                ),
            ]
        elif plat == "facebook":
            rows_pm = [("Views", met.get("views", 0)), ("Clicks", met.get("clicks", 0))]
        elif plat == "instagram":
            eng_top = met.get("likes", 0) + met.get("comments", 0) + met.get("shares", 0)
            rows_pm = [
                ("Views", met.get("views", 0)),
                ("Reach", met.get("reach", 0)),
                ("Engagement", eng_top),
            ]
        elif plat == "youtube":
            rows_pm = [
                ("Views", met.get("views", 0)),
                ("Likes", met.get("likes", 0)),
                ("Shares", met.get("shares", 0)),
            ]
        elif plat == "twitter":
            rows_pm = [
                ("Impressions", met.get("impressions", 0)),
                ("Likes", met.get("likes", 0)),
                ("Reposts", met.get("reposts", 0)),
            ]
        else:
            rows_pm = [
                ("Views", met.get("views", 0)),
                ("Likes", met.get("likes", 0)),
                ("Comments", met.get("comments", 0)),
                ("Shares", met.get("shares", 0)),
            ]

        for label, val in rows_pm:
            ws.cell(row=pr, column=c0, value=label).font = Font(bold=True)
            ws.merge_cells(start_row=pr, start_column=c0 + 1, end_row=pr, end_column=c0 + 3)
            ws.cell(row=pr, column=c0 + 1, value=val).alignment = Alignment(horizontal="left", wrap_text=True)
            pr += 1

        # Thumbnail
        img_url = None
        if top_h and top_h.get("mediaUrls"):
            img_url = top_h["mediaUrls"][0]
        if img_url and re.search(r"\.(jpe?g|png|gif|webp)(\?|$)", str(img_url), re.I):
            try:
                ir = requests.get(str(img_url), timeout=25)
                if ir.ok and ir.content:
                    bio = io.BytesIO(ir.content)
                    img = XLImage(bio)
                    img.width = min(img.width, 220)
                    img.height = min(img.height, 180)
                    anchor = f"{get_column_letter(c0 + 1)}{pr}"
                    ws.add_image(img, anchor)
                    ws.row_dimensions[pr].height = 120
            except Exception:
                ws.cell(row=pr, column=c0, value="Image").font = Font(bold=True)
                ws.cell(
                    row=pr, column=c0 + 1, value="(Preview unavailable — video or blocked URL)"
                ).alignment = Alignment(wrap_text=True)
        else:
            ws.cell(row=pr, column=c0, value="Image").font = Font(bold=True)
            ws.cell(row=pr, column=c0 + 1, value="(No static image URL — check post link)").alignment = (
                Alignment(wrap_text=True)
            )

    for i in range(len(PLATFORMS)):
        c0 = block_start(i)
        for cc in range(c0, c0 + 5):
            ws.column_dimensions[get_column_letter(cc)].width = 14 if cc > c0 else 22

    wb.save(OUT_XLSX)
    print(f"Wrote {OUT_XLSX}")


# ── Main ───────────────────────────────────────────────────────────────────────


def main() -> None:
    os.makedirs(OUT_JSON_DIR, exist_ok=True)
    sess = _session()

    start_a, end_a = month_range(YEAR, MONTH_A)
    start_b, end_b = month_range(YEAR, MONTH_B)

    print(f"Fetching history {MONTH_LABELS[0]}…")
    hist_a = fetch_history(sess, start_a, end_a)
    print(f"Fetching history {MONTH_LABELS[1]}…")
    hist_b = fetch_history(sess, start_b, end_b)

    cache_dir = OUT_JSON_DIR
    hist_a_path = os.path.join(cache_dir, "history_feb.json")
    hist_b_path = os.path.join(cache_dir, "history_mar.json")
    social_a_path = os.path.join(cache_dir, "social_feb.json")
    social_b_path = os.path.join(cache_dir, "social_mar.json")

    with open(hist_a_path, "w") as f:
        json.dump(hist_a, f, indent=2)
    with open(hist_b_path, "w") as f:
        json.dump(hist_b, f, indent=2)

    print(f"Fetching social {MONTH_LABELS[0]}…")
    social_a = fetch_social(sess, start_a, end_a)
    print(f"Fetching social {MONTH_LABELS[1]}…")
    social_b = fetch_social(sess, start_b, end_b)

    with open(social_a_path, "w") as f:
        json.dump(social_a, f, indent=2)
    with open(social_b_path, "w") as f:
        json.dump(social_b, f, indent=2)

    by_a = collect_posts_for_month(hist_a, YEAR, MONTH_A)
    by_b = collect_posts_for_month(hist_b, YEAR, MONTH_B)

    jobs: list[tuple[str, str]] = []
    for plat in PLATFORMS:
        for h in by_a[plat] + by_b[plat]:
            jobs.append((str(h["id"]), plat))
    jobs = list(dict.fromkeys(jobs))

    print(f"Fetching post analytics: {len(jobs)} calls (workers={MAX_WORKERS})…")
    post_metrics: dict[tuple[str, str], dict[str, float]] = {}
    raw_dump: list[dict[str, Any]] = []

    def one(job: tuple[str, str]) -> tuple[tuple[str, str], dict[str, float], dict[str, Any]]:
        pid, plat = job
        last_err = ""
        for attempt in range(MAX_RETRIES):
            try:
                body = fetch_post_analytics(sess, pid, plat)
                raw_dump.append({"post_id": pid, "platform": plat, "attempt": attempt + 1, "body": body})
                return (job, extract_post_metrics(plat, body), body)
            except Exception as e:
                last_err = str(e)
                time.sleep(1.4**attempt)
        return (job, extract_post_metrics(plat, {"status": "error"}), {"error": last_err})

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
        futs = [ex.submit(one, j) for j in jobs]
        for k, fut in enumerate(as_completed(futs), 1):
            job, metrics, _body = fut.result()
            post_metrics[job] = metrics
            if k % 10 == 0 or k == len(futs):
                print(f"  … {k}/{len(futs)}")

    with open(os.path.join(OUT_JSON_DIR, "post_analytics_merged.json"), "w") as f:
        json.dump(raw_dump, f, indent=2)

    notes = (
        f"Generated {datetime.now().isoformat(timespec='seconds')} — live Ayrshare scrape. "
        f"Columns: {MONTH_LABELS[0]} vs {MONTH_LABELS[1]}. "
        f"MoM % Δ uses ({MONTH_LABELS[1]}−{MONTH_LABELS[0]})/{MONTH_LABELS[0]}×100 when {MONTH_LABELS[0]}≠0; otherwise '—'. "
        f"Follower Growth % row shows the monthly change; Jan col = '—'. "
        f"Post-level metrics summed across all posts in the period; fallback to social-API period numbers if zero."
    )

    build_excel(by_a, by_b, social_a, social_b, post_metrics, notes)
    print("Done.")


if __name__ == "__main__":
    main()
